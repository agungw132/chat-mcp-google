import sys

import pytest

from chat_google.mcp_servers import sheets_server


def _build_excel_bytes():
    try:
        from openpyxl import Workbook
    except Exception:
        pytest.skip("openpyxl not available")
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "SheetA"
    ws["A1"] = "name"
    ws["B1"] = "value"
    ws["A2"] = "ops"
    ws["B2"] = "42"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_sheets_tools_smoke(monkeypatch):
    async def fake_drive_get(path, params=None):
        if path == "/files" and "mimeType='application/vnd.google-apps.folder'" in (params or {}).get("q", ""):
            return ({"files": [{"id": "folder-docs", "name": "Documents"}]}, None)
        if path == "/files/tpl1":
            return (
                {
                    "id": "tpl1",
                    "name": "Template Sheet",
                    "mimeType": "application/vnd.google-apps.spreadsheet",
                    "parents": ["parent-1"],
                    "webViewLink": "https://docs.google.com/spreadsheets/d/tpl1/edit",
                },
                None,
            )
        if path == "/files/sp-new-template":
            return (
                {
                    "id": "sp-new-template",
                    "name": "Budget From Template",
                    "mimeType": "application/vnd.google-apps.spreadsheet",
                    "parents": ["parent-1"],
                    "webViewLink": "https://docs.google.com/spreadsheets/d/sp-new-template/edit",
                },
                None,
            )
        if path == "/files/sp1":
            return (
                {
                    "id": "sp1",
                    "name": "Budget",
                    "mimeType": "application/vnd.google-apps.spreadsheet",
                    "modifiedTime": "2026-02-14T08:00:00Z",
                    "size": "45678",
                    "webViewLink": "https://docs.google.com/spreadsheets/d/sp1/edit",
                    "owners": [{"displayName": "Owner", "emailAddress": "owner@example.com"}],
                    "parents": ["parent-1"],
                },
                None,
            )
        if path == "/files/sp1/permissions":
            return (
                {
                    "permissions": [
                        {
                            "id": "perm-1",
                            "type": "user",
                            "role": "writer",
                            "emailAddress": "alice@example.com",
                        }
                    ]
                },
                None,
            )
        if path == "/files/x1":
            return (
                {
                    "id": "x1",
                    "name": "Ops.xlsx",
                    "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "modifiedTime": "2026-02-13T07:00:00Z",
                    "size": "12345",
                    "webViewLink": "https://drive.google.com/file/d/x1/view",
                    "owners": [{"displayName": "Owner", "emailAddress": "owner@example.com"}],
                    "parents": ["parent-1"],
                },
                None,
            )
        if path == "/files/xl1":
            return (
                {
                    "id": "xl1",
                    "name": "Legacy.xls",
                    "mimeType": "application/vnd.ms-excel",
                    "modifiedTime": "2026-02-12T07:00:00Z",
                    "size": "9876",
                    "webViewLink": "https://drive.google.com/file/d/xl1/view",
                    "owners": [{"displayName": "Owner", "emailAddress": "owner@example.com"}],
                    "parents": ["parent-1"],
                },
                None,
            )
        if path == "/files" and "name contains" not in (params or {}).get("q", ""):
            return (
                {
                    "files": [
                        {
                            "id": "sp1",
                            "name": "Budget",
                            "mimeType": "application/vnd.google-apps.spreadsheet",
                            "modifiedTime": "2026-02-14T08:00:00Z",
                            "webViewLink": "https://docs.google.com/spreadsheets/d/sp1/edit",
                        },
                        {
                            "id": "x1",
                            "name": "Ops.xlsx",
                            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            "modifiedTime": "2026-02-13T07:00:00Z",
                            "webViewLink": "https://drive.google.com/file/d/x1/view",
                        },
                        {
                            "id": "xl1",
                            "name": "Legacy.xls",
                            "mimeType": "application/vnd.ms-excel",
                            "modifiedTime": "2026-02-12T07:00:00Z",
                            "webViewLink": "https://drive.google.com/file/d/xl1/view",
                        },
                    ]
                },
                None,
            )
        if path == "/files" and "name contains" in (params or {}).get("q", ""):
            return (
                {
                    "files": [
                        {
                            "id": "sp1",
                            "name": "Budget",
                            "mimeType": "application/vnd.google-apps.spreadsheet",
                            "modifiedTime": "2026-02-14T08:00:00Z",
                            "webViewLink": "https://docs.google.com/spreadsheets/d/sp1/edit",
                        }
                    ]
                },
                None,
            )
        return {}, None

    async def fake_sheets_get(path, params=None):
        if path == "/sp1":
            return (
                {
                    "spreadsheetId": "sp1",
                    "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp1/edit",
                    "properties": {"title": "Budget"},
                    "sheets": [{"properties": {"sheetId": 0, "title": "Sheet1", "gridProperties": {"rowCount": 1000, "columnCount": 26}}}],
                },
                None,
            )
        if path == "/sp1/values:batchGet":
            return (
                {
                    "valueRanges": [
                        {"range": "Sheet1!A1:B2", "values": [["name", "value"], ["ops", "42"]]},
                        {"range": "Sheet1!D1:D2", "values": [["status"], ["active"]]},
                    ]
                },
                None,
            )
        if "/values/" in path:
            return {"range": "Sheet1!A1:B2", "values": [["name", "value"], ["ops", "42"]]}, None
        return {}, None

    async def fake_drive_get_bytes(path, params=None):
        if path == "/files/xl1":
            return b"legacy-xls", None
        return _build_excel_bytes(), None

    async def fake_drive_post(path, params=None, json_body=None):
        if path == "/files/x1/copy":
            return {"id": "sp-conv", "name": "Ops"}, None
        if path == "/files/sp1/permissions":
            return {"id": "perm-1"}, None
        if path == "/files/tpl1/copy":
            return {"id": "sp-new-template", "name": "Budget From Template"}, None
        return {}, None

    async def fake_sheets_post(path, json_body=None, params=None):
        if path.endswith(":append"):
            return {"updates": {"updatedRange": "Sheet1!A3:B3", "updatedRows": 1, "updatedCells": 2}}, None
        if path.endswith(":clear"):
            return {"clearedRange": "Sheet1!A:ZZZ"}, None
        if path == "":
            return (
                {
                    "spreadsheetId": "sp-new",
                    "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp-new/edit",
                    "properties": {"title": "Roadmap"},
                    "sheets": [{"properties": {"sheetId": 0, "title": "Plan"}}],
                },
                None,
            )
        if path == "/sp1:batchUpdate":
            request = (json_body or {}).get("requests", [])[0]
            if "addSheet" in request:
                return {"replies": [{"addSheet": {"properties": {"sheetId": 2, "title": "Q2"}}}]}, None
            if "addChart" in request:
                return {"replies": [{"addChart": {"chart": {"chartId": 88}}}]}, None
            if "addProtectedRange" in request:
                return {"replies": [{"addProtectedRange": {"protectedRange": {"protectedRangeId": 77}}}]}, None
            if "updateCells" in request:
                return {"replies": [{}]}, None
        if path == "/sp1/values:batchUpdate":
            return (
                {
                    "totalUpdatedRows": 3,
                    "totalUpdatedColumns": 4,
                    "totalUpdatedCells": 10,
                    "responses": [
                        {"updatedRange": "Sheet1!A1:B2", "updatedRows": 2, "updatedColumns": 2, "updatedCells": 4},
                        {"updatedRange": "Sheet1!D1:D2", "updatedRows": 2, "updatedColumns": 1, "updatedCells": 2},
                    ],
                },
                None,
            )
        return {}, None

    async def fake_sheets_put(path, json_body=None, params=None):
        return {
            "updatedRange": "Sheet1!A1:B1",
            "updatedRows": 1,
            "updatedColumns": 2,
            "updatedCells": 2,
        }, None

    async def fake_drive_export(file_id, export_format, mime_type, gid=None, range_a1=""):
        return b"name,value\nops,42\n", None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_get_bytes", fake_drive_get_bytes)
    monkeypatch.setattr(sheets_server, "_drive_post", fake_drive_post)
    monkeypatch.setattr(sheets_server, "_drive_export_google_sheet_bytes", fake_drive_export)
    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    monkeypatch.setattr(sheets_server, "_sheets_put", fake_sheets_put)

    class _FakeXlsSheet:
        nrows = 2
        ncols = 2

        @staticmethod
        def cell_value(row_idx, col_idx):
            values = [
                ["name", "value"],
                ["legacy", 7.0],
            ]
            return values[row_idx][col_idx]

    class _FakeXlsWorkbook:
        @staticmethod
        def sheet_names():
            return ["SheetA"]

        @staticmethod
        def sheet_by_name(name):
            assert name == "SheetA"
            return _FakeXlsSheet()

    class _FakeXlrd:
        @staticmethod
        def open_workbook(file_contents=None):
            assert file_contents == b"legacy-xls"
            return _FakeXlsWorkbook()

    monkeypatch.setitem(sys.modules, "xlrd", _FakeXlrd())

    listed = await sheets_server.list_sheets_spreadsheets()
    searched = await sheets_server.search_sheets_spreadsheets("Budget")
    metadata = await sheets_server.get_sheets_metadata("sp1")
    unified_metadata = await sheets_server.get_spreadsheet_metadata("x1")
    unified_metadata_xls = await sheets_server.get_spreadsheet_metadata("xl1")
    read = await sheets_server.read_sheet_values("sp1", "Sheet1!A1:B2")
    read_excel = await sheets_server.read_excel_values("x1", range_a1="A1:B2")
    read_excel_xls = await sheets_server.read_excel_values("xl1", range_a1="A1:B2")
    appended = await sheets_server.append_sheet_row("sp1", "Sheet1!A:B", ["ops", "42"])
    updated = await sheets_server.update_sheet_values("sp1", "Sheet1!A1:B1", [["a", "b"]])
    created = await sheets_server.create_sheets_spreadsheet("Roadmap", sheet_title="Plan")
    added_tab = await sheets_server.add_sheet_tab("sp1", "Q2")
    listed_unified = await sheets_server.list_spreadsheets(limit=3, include_excel=True)
    template_list = await sheets_server.list_spreadsheet_templates(limit=3, folder_name="Documents")
    searched_unified = await sheets_server.search_spreadsheets("Budget", include_excel=True)
    converted = await sheets_server.convert_excel_to_google_sheet("x1")
    exported = await sheets_server.export_google_sheet("sp1", export_format="csv")
    batch_read = await sheets_server.batch_get_sheet_values(
        "sp1",
        ["Sheet1!A1:B2", "Sheet1!D1:D2"],
    )
    batch_update = await sheets_server.batch_update_sheet_values(
        "sp1",
        updates=[
            {"range_a1": "Sheet1!A1:B2", "values": [["name", "value"], ["ops", "42"]]},
            {"range_a1": "Sheet1!D1:D2", "values": [["status"], ["active"]]},
        ],
    )
    shared = await sheets_server.share_spreadsheet("sp1", "alice@example.com")
    template_copied = await sheets_server.create_spreadsheet_from_template(
        "tpl1",
        "Budget From Template",
    )
    csv_imported = await sheets_server.import_csv_to_sheet(
        "sp1",
        "Sheet1",
        "name,value\nops,42\n",
        overwrite=True,
    )
    chart_inserted = await sheets_server.insert_sheet_chart(
        "sp1",
        0,
        {"spec": {"title": "Ops", "basicChart": {"chartType": "COLUMN"}}},
    )
    protected = await sheets_server.protect_sheet_or_range(
        "sp1",
        sheet_id=0,
        editors=["alice@example.com"],
    )
    pivoted = await sheets_server.create_pivot_table(
        "sp1",
        source_range="Sheet1!A1:B10",
        target_sheet="Sheet1",
        target_cell="D1",
    )
    permissions = await sheets_server.get_spreadsheet_permissions("sp1")

    assert "Google Sheets files" in listed
    assert "Google Sheets search results" in searched
    assert "Google Sheets Metadata" in metadata
    assert "Spreadsheet Metadata" in unified_metadata
    assert "excel_xls" in unified_metadata_xls
    assert "Google Sheets Values" in read
    assert "Excel (.xlsx) Values" in read_excel
    assert "Excel (.xls) Values" in read_excel_xls
    assert "Google Sheets append row completed" in appended
    assert "Google Sheets update values completed" in updated
    assert "Google Sheets spreadsheet created" in created
    assert "Google Sheets tab created" in added_tab
    assert "Spreadsheet files (include_excel=True" in listed_unified
    assert "excel_xls" in listed_unified
    assert "Spreadsheet templates in folder 'Documents'" in template_list
    assert "Spreadsheet search results for 'Budget'" in searched_unified
    assert "conversion completed" in converted.lower()
    assert "Google Sheets export completed" in exported
    assert "Google Sheets batch get values completed" in batch_read
    assert "Google Sheets batch update completed" in batch_update
    assert "Spreadsheet share completed" in shared
    assert "Spreadsheet created from template" in template_copied
    assert "CSV import to sheet completed" in csv_imported
    assert "Sheet chart inserted" in chart_inserted
    assert "Sheet protection created" in protected
    assert "Pivot table created" in pivoted
    assert "Spreadsheet permissions" in permissions
