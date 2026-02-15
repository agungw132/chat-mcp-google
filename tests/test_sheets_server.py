import sys

import pytest

from chat_google.mcp_servers import sheets_server


def _build_excel_bytes():
    try:
        from openpyxl import Workbook
    except Exception:
        pytest.skip("openpyxl not available")
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "SheetA"
    ws["A1"] = "name"
    ws["B1"] = "value"
    ws["A2"] = "ops"
    ws["B2"] = 42
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_get_access_token_missing(monkeypatch):
    monkeypatch.setattr(sheets_server, "load_dotenv", lambda *args, **kwargs: None)
    monkeypatch.setattr(sheets_server, "_CACHED_ACCESS_TOKEN", None)
    monkeypatch.setattr(sheets_server, "_CACHED_ACCESS_TOKEN_EXPIRES_AT", None)
    monkeypatch.delenv("GOOGLE_DRIVE_ACCESS_TOKEN", raising=False)
    monkeypatch.delenv("GOOGLE_DRIVE_REFRESH_TOKEN", raising=False)
    monkeypatch.delenv("GOOGLE_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("GOOGLE_OAUTH_CLIENT_SECRET", raising=False)
    with pytest.raises(ValueError):
        sheets_server._get_access_token()


def test_get_access_token_uses_refresh_flow(monkeypatch):
    monkeypatch.setattr(sheets_server, "load_dotenv", lambda *args, **kwargs: None)
    monkeypatch.setattr(sheets_server, "_CACHED_ACCESS_TOKEN", None)
    monkeypatch.setattr(sheets_server, "_CACHED_ACCESS_TOKEN_EXPIRES_AT", None)
    monkeypatch.delenv("GOOGLE_DRIVE_ACCESS_TOKEN", raising=False)
    monkeypatch.setenv("GOOGLE_DRIVE_REFRESH_TOKEN", "refresh-token")
    monkeypatch.setenv("GOOGLE_OAUTH_CLIENT_ID", "client-id")
    monkeypatch.setenv("GOOGLE_OAUTH_CLIENT_SECRET", "client-secret")

    def fake_refresh_access_token(refresh_token, client_id, client_secret):
        assert refresh_token == "refresh-token"
        assert client_id == "client-id"
        assert client_secret == "client-secret"
        return "refreshed-access-token", 3600

    monkeypatch.setattr(sheets_server, "_refresh_access_token", fake_refresh_access_token)
    token = sheets_server._get_access_token()
    assert token == "refreshed-access-token"


@pytest.mark.asyncio
async def test_list_sheets_spreadsheets(monkeypatch):
    async def fake_drive_get(path, params=None):
        assert path == "/files"
        return (
            {
                "files": [
                    {
                        "id": "sheet1",
                        "name": "Budget",
                        "modifiedTime": "2026-02-14T09:00:00Z",
                        "webViewLink": "https://docs.google.com/spreadsheets/d/sheet1/edit",
                    }
                ]
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    result = await sheets_server.list_sheets_spreadsheets(limit=1)
    assert "Google Sheets files (showing 1):" in result
    assert "Budget" in result
    assert "sheet1" in result


@pytest.mark.asyncio
async def test_search_sheets_spreadsheets_no_results(monkeypatch):
    async def fake_drive_get(path, params=None):
        return {"files": []}, None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    result = await sheets_server.search_sheets_spreadsheets("Quarterly")
    assert result == "No Google Sheets files found matching 'Quarterly'"


@pytest.mark.asyncio
async def test_get_sheets_metadata(monkeypatch):
    async def fake_sheets_get(path, params=None):
        assert path == "/sp123"
        return (
            {
                "spreadsheetId": "sp123",
                "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp123/edit",
                "properties": {"title": "Finance"},
                "sheets": [
                    {
                        "properties": {
                            "sheetId": 0,
                            "title": "Jan",
                            "gridProperties": {"rowCount": 1000, "columnCount": 26},
                        }
                    }
                ],
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    result = await sheets_server.get_sheets_metadata("sp123")
    assert "Google Sheets Metadata:" in result
    assert "Title: Finance" in result
    assert "Jan" in result


@pytest.mark.asyncio
async def test_read_sheet_values_limits(monkeypatch):
    async def fake_sheets_get(path, params=None):
        assert "/values/" in path
        return (
            {
                "range": "Sheet1!A1:C3",
                "values": [
                    ["h1", "h2", "h3"],
                    ["r1c1", "r1c2", "r1c3"],
                    ["r2c1", "r2c2", "r2c3"],
                ],
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    result = await sheets_server.read_sheet_values("sp123", "Sheet1!A1:C3", max_rows=2, max_cols=2)
    assert "Rows Returned: 3" in result
    assert "Rows Shown: 2" in result
    assert "Row display limited to first 2 rows." in result
    assert "Column display limited to first 2 columns per row." in result


@pytest.mark.asyncio
async def test_append_sheet_row(monkeypatch):
    async def fake_sheets_post(path, json_body=None, params=None):
        assert path.endswith(":append")
        assert json_body["values"] == [["100", "200"]]
        return {"updates": {"updatedRange": "Sheet1!A10:B10", "updatedRows": 1, "updatedCells": 2}}, None

    async def fake_sheets_get(path, params=None):
        return {"spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp123/edit", "properties": {"title": "Budget"}}, None

    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    result = await sheets_server.append_sheet_row("sp123", "Sheet1!A:B", ["100", "200"])
    assert "Google Sheets append row completed:" in result
    assert "Updated Cells: 2" in result


@pytest.mark.asyncio
async def test_update_sheet_values(monkeypatch):
    async def fake_sheets_put(path, json_body=None, params=None):
        assert "/values/" in path
        assert params["valueInputOption"] == "RAW"
        return {
            "updatedRange": "Sheet1!A1:B1",
            "updatedRows": 1,
            "updatedColumns": 2,
            "updatedCells": 2,
        }, None

    monkeypatch.setattr(sheets_server, "_sheets_put", fake_sheets_put)
    result = await sheets_server.update_sheet_values(
        "sp123",
        "Sheet1!A1:B1",
        [["x", "y"]],
        value_input_option="RAW",
    )
    assert "Google Sheets update values completed:" in result
    assert "Updated Cells: 2" in result
    assert "Value Input Option: RAW" in result


@pytest.mark.asyncio
async def test_create_sheets_spreadsheet(monkeypatch):
    async def fake_sheets_post(path, json_body=None, params=None):
        assert path == ""
        assert json_body["properties"]["title"] == "Roadmap"
        return (
            {
                "spreadsheetId": "sp-new",
                "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp-new/edit",
                "properties": {"title": "Roadmap"},
                "sheets": [{"properties": {"sheetId": 0, "title": "Plan"}}],
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    result = await sheets_server.create_sheets_spreadsheet("Roadmap", sheet_title="Plan")
    assert "Google Sheets spreadsheet created:" in result
    assert "Spreadsheet ID: sp-new" in result


@pytest.mark.asyncio
async def test_add_sheet_tab(monkeypatch):
    async def fake_sheets_post(path, json_body=None, params=None):
        assert path == "/sp123:batchUpdate"
        return {"replies": [{"addSheet": {"properties": {"sheetId": 7, "title": "Q2"}}}]}, None

    async def fake_sheets_get(path, params=None):
        return {"spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp123/edit", "properties": {"title": "Budget"}}, None

    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    result = await sheets_server.add_sheet_tab("sp123", "Q2", row_count=200, column_count=30)
    assert "Google Sheets tab created:" in result
    assert "Tab ID: 7" in result


@pytest.mark.asyncio
async def test_read_sheet_values_propagates_error(monkeypatch):
    async def fake_sheets_get(path, params=None):
        return None, "Error: Google Sheets API request failed: 403 - forbidden"

    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    result = await sheets_server.read_sheet_values("sp123", "A1:B2")
    assert result == "Error: Google Sheets API request failed: 403 - forbidden"


@pytest.mark.asyncio
async def test_list_spreadsheets_include_excel(monkeypatch):
    async def fake_drive_get(path, params=None):
        assert "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'" in params["q"]
        assert "mimeType='application/vnd.ms-excel'" in params["q"]
        return (
            {
                "files": [
                    {
                        "id": "sp1",
                        "name": "Native Sheet",
                        "mimeType": "application/vnd.google-apps.spreadsheet",
                        "modifiedTime": "2026-02-14T09:00:00Z",
                        "webViewLink": "https://docs.google.com/spreadsheets/d/sp1/edit",
                    },
                    {
                        "id": "x1",
                        "name": "Excel File.xlsx",
                        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "modifiedTime": "2026-02-13T09:00:00Z",
                        "webViewLink": "https://drive.google.com/file/d/x1/view",
                    },
                    {
                        "id": "xl1",
                        "name": "Legacy File.xls",
                        "mimeType": "application/vnd.ms-excel",
                        "modifiedTime": "2026-02-12T09:00:00Z",
                        "webViewLink": "https://drive.google.com/file/d/xl1/view",
                    },
                ]
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    result = await sheets_server.list_spreadsheets(limit=2, include_excel=True)
    assert "include_excel=True" in result
    assert "Type: google_sheet" in result
    assert "Type: excel_xlsx" in result
    assert "Type: excel_xls" in result


@pytest.mark.asyncio
async def test_get_spreadsheet_metadata_excel(monkeypatch):
    async def fake_drive_get(path, params=None):
        assert path == "/files/x1"
        return (
            {
                "id": "x1",
                "name": "Excel File.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-02-13T09:00:00Z",
                "size": "1234",
                "webViewLink": "https://drive.google.com/file/d/x1/view",
                "owners": [{"displayName": "Alice", "emailAddress": "alice@example.com"}],
            },
            None,
        )

    async def fake_drive_get_bytes(path, params=None):
        return _build_excel_bytes(), None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_get_bytes", fake_drive_get_bytes)
    result = await sheets_server.get_spreadsheet_metadata("x1")
    assert "Type: excel_xlsx" in result
    assert "SheetA" in result


@pytest.mark.asyncio
async def test_get_spreadsheet_metadata_excel_xls(monkeypatch):
    async def fake_drive_get(path, params=None):
        assert path == "/files/xl1"
        return (
            {
                "id": "xl1",
                "name": "Legacy File.xls",
                "mimeType": "application/vnd.ms-excel",
                "modifiedTime": "2026-02-13T09:00:00Z",
                "size": "4321",
                "webViewLink": "https://drive.google.com/file/d/xl1/view",
                "owners": [{"displayName": "Alice", "emailAddress": "alice@example.com"}],
            },
            None,
        )

    async def fake_drive_get_bytes(path, params=None):
        return b"legacy-xls", None

    class _FakeXlsWorkbook:
        @staticmethod
        def sheet_names():
            return ["LegacySheet"]

    class _FakeXlrd:
        @staticmethod
        def open_workbook(file_contents=None):
            assert file_contents == b"legacy-xls"
            return _FakeXlsWorkbook()

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_get_bytes", fake_drive_get_bytes)
    monkeypatch.setitem(sys.modules, "xlrd", _FakeXlrd())
    result = await sheets_server.get_spreadsheet_metadata("xl1")
    assert "Type: excel_xls" in result
    assert "LegacySheet" in result


@pytest.mark.asyncio
async def test_read_excel_values(monkeypatch):
    async def fake_drive_get(path, params=None):
        return (
            {
                "id": "x1",
                "name": "Excel File.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "webViewLink": "https://drive.google.com/file/d/x1/view",
            },
            None,
        )

    async def fake_drive_get_bytes(path, params=None):
        return _build_excel_bytes(), None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_get_bytes", fake_drive_get_bytes)
    result = await sheets_server.read_excel_values("x1", range_a1="A1:B2", max_rows=10, max_cols=10)
    assert "Excel (.xlsx) Values:" in result
    assert "name | value" in result
    assert "ops | 42" in result


@pytest.mark.asyncio
async def test_read_excel_values_xls(monkeypatch):
    async def fake_drive_get(path, params=None):
        return (
            {
                "id": "xl1",
                "name": "Legacy File.xls",
                "mimeType": "application/vnd.ms-excel",
                "webViewLink": "https://drive.google.com/file/d/xl1/view",
            },
            None,
        )

    async def fake_drive_get_bytes(path, params=None):
        return b"legacy-xls", None

    class _FakeXlsSheet:
        nrows = 2
        ncols = 2

        @staticmethod
        def cell_value(row_idx, col_idx):
            values = [
                ["name", "value"],
                ["ops", 42.0],
            ]
            return values[row_idx][col_idx]

    class _FakeXlsWorkbook:
        @staticmethod
        def sheet_names():
            return ["SheetA"]

        @staticmethod
        def sheet_by_name(name):
            assert name == "SheetA"
            return _FakeXlsSheet()

    class _FakeXlrd:
        @staticmethod
        def open_workbook(file_contents=None):
            assert file_contents == b"legacy-xls"
            return _FakeXlsWorkbook()

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_get_bytes", fake_drive_get_bytes)
    monkeypatch.setitem(sys.modules, "xlrd", _FakeXlrd())
    result = await sheets_server.read_excel_values("xl1", range_a1="A1:B2", max_rows=10, max_cols=10)
    assert "Excel (.xls) Values:" in result
    assert "name | value" in result
    assert "ops | 42" in result


@pytest.mark.asyncio
async def test_convert_excel_to_google_sheet_copy_success(monkeypatch):
    async def fake_drive_get(path, params=None):
        return (
            {
                "id": "x1",
                "name": "Excel File.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "parents": ["p1"],
            },
            None,
        )

    async def fake_drive_post(path, params=None, json_body=None):
        assert path == "/files/x1/copy"
        assert json_body["mimeType"] == "application/vnd.google-apps.spreadsheet"
        return {"id": "sp-new", "name": "Excel File"}, None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_post", fake_drive_post)
    result = await sheets_server.convert_excel_to_google_sheet("x1")
    assert "conversion completed" in result.lower()
    assert "Spreadsheet ID: sp-new" in result


@pytest.mark.asyncio
async def test_convert_excel_to_google_sheet_copy_success_xls(monkeypatch):
    async def fake_drive_get(path, params=None):
        return (
            {
                "id": "xl1",
                "name": "Excel File.xls",
                "mimeType": "application/vnd.ms-excel",
                "parents": ["p1"],
            },
            None,
        )

    async def fake_drive_post(path, params=None, json_body=None):
        assert path == "/files/xl1/copy"
        assert json_body["mimeType"] == "application/vnd.google-apps.spreadsheet"
        return {"id": "sp-new", "name": "Excel File"}, None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_post", fake_drive_post)
    result = await sheets_server.convert_excel_to_google_sheet("xl1")
    assert "conversion completed" in result.lower()
    assert "Spreadsheet ID: sp-new" in result


@pytest.mark.asyncio
async def test_export_google_sheet_csv_preview(monkeypatch):
    async def fake_drive_get(path, params=None):
        assert path == "/files/sp1"
        return (
            {
                "id": "sp1",
                "name": "Budget",
                "mimeType": "application/vnd.google-apps.spreadsheet",
                "webViewLink": "https://docs.google.com/spreadsheets/d/sp1/edit",
            },
            None,
        )

    async def fake_drive_export(file_id, export_format, mime_type, gid=None, range_a1=""):
        assert file_id == "sp1"
        assert export_format == "csv"
        assert mime_type == "text/csv"
        return b"name,value\nops,42\n", None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_export_google_sheet_bytes", fake_drive_export)
    result = await sheets_server.export_google_sheet("sp1", export_format="csv")
    assert "Google Sheets export completed:" in result
    assert "Format: csv" in result
    assert "name,value" in result


@pytest.mark.asyncio
async def test_export_google_sheet_requires_native_sheet(monkeypatch):
    async def fake_drive_get(path, params=None):
        return (
            {
                "id": "x1",
                "name": "Budget.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    result = await sheets_server.export_google_sheet("x1", export_format="xlsx")
    assert "Use convert_excel_to_google_sheet" in result


@pytest.mark.asyncio
async def test_batch_get_sheet_values(monkeypatch):
    async def fake_sheets_get(path, params=None):
        assert path == "/sp1/values:batchGet"
        assert params["ranges"] == ["Sheet1!A1:B2", "Sheet2!A1:A2"]
        return (
            {
                "valueRanges": [
                    {"range": "Sheet1!A1:B2", "values": [["name", "value"], ["ops", "42"]]},
                    {"range": "Sheet2!A1:A2", "values": [["team"], ["platform"]]},
                ]
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    result = await sheets_server.batch_get_sheet_values(
        "sp1",
        ["Sheet1!A1:B2", "Sheet2!A1:A2"],
        max_rows_per_range=5,
        max_cols_per_row=5,
    )
    assert "Google Sheets batch get values completed:" in result
    assert "Ranges Returned: 2" in result
    assert "Sheet1!A1:B2" in result


@pytest.mark.asyncio
async def test_batch_update_sheet_values(monkeypatch):
    async def fake_sheets_post(path, json_body=None, params=None):
        assert path == "/sp1/values:batchUpdate"
        assert json_body["valueInputOption"] == "RAW"
        assert len(json_body["data"]) == 2
        return (
            {
                "totalUpdatedRows": 3,
                "totalUpdatedColumns": 4,
                "totalUpdatedCells": 10,
                "responses": [
                    {"updatedRange": "Sheet1!A1:B2", "updatedRows": 2, "updatedColumns": 2, "updatedCells": 4},
                    {"updatedRange": "Sheet2!A1:B1", "updatedRows": 1, "updatedColumns": 2, "updatedCells": 2},
                ],
            },
            None,
        )

    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    result = await sheets_server.batch_update_sheet_values(
        "sp1",
        updates=[
            {"range_a1": "Sheet1!A1:B2", "values": [["a", "b"], ["c", "d"]]},
            {"range_a1": "Sheet2!A1:B1", "values": [["x", "y"]]},
        ],
        value_input_option="RAW",
    )
    assert "Google Sheets batch update completed:" in result
    assert "Total Updated Cells: 10" in result
    assert "Update 1: Sheet1!A1:B2" in result


@pytest.mark.asyncio
async def test_share_spreadsheet(monkeypatch):
    async def fake_drive_get(path, params=None):
        assert path == "/files/sp1"
        return (
            {
                "id": "sp1",
                "name": "Budget",
                "mimeType": "application/vnd.google-apps.spreadsheet",
                "webViewLink": "https://docs.google.com/spreadsheets/d/sp1/edit",
            },
            None,
        )

    async def fake_drive_post(path, params=None, json_body=None):
        assert path == "/files/sp1/permissions"
        assert json_body["emailAddress"] == "alice@example.com"
        assert json_body["role"] == "writer"
        return {"id": "perm-1"}, None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_post", fake_drive_post)
    result = await sheets_server.share_spreadsheet(
        "sp1",
        "alice@example.com",
        role="writer",
        send_notification=True,
    )
    assert "Spreadsheet share completed:" in result
    assert "Permission ID: perm-1" in result
    assert "alice@example.com" in result


@pytest.mark.asyncio
async def test_share_spreadsheet_xls(monkeypatch):
    async def fake_drive_get(path, params=None):
        assert path == "/files/xl1"
        return (
            {
                "id": "xl1",
                "name": "Legacy.xls",
                "mimeType": "application/vnd.ms-excel",
                "webViewLink": "https://drive.google.com/file/d/xl1/view",
            },
            None,
        )

    async def fake_drive_post(path, params=None, json_body=None):
        assert path == "/files/xl1/permissions"
        assert json_body["emailAddress"] == "alice@example.com"
        assert json_body["role"] == "reader"
        return {"id": "perm-xls"}, None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_post", fake_drive_post)
    result = await sheets_server.share_spreadsheet(
        "xl1",
        "alice@example.com",
        role="reader",
        send_notification=False,
    )
    assert "Spreadsheet share completed:" in result
    assert "Type: excel_xls" in result
    assert "Permission ID: perm-xls" in result


@pytest.mark.asyncio
async def test_create_spreadsheet_from_template(monkeypatch):
    async def fake_drive_get(path, params=None):
        if path == "/files/tpl1":
            return (
                {
                    "id": "tpl1",
                    "name": "Template Budget",
                    "mimeType": "application/vnd.google-apps.spreadsheet",
                    "parents": ["folder-a"],
                },
                None,
            )
        if path == "/files/new1":
            return (
                {
                    "id": "new1",
                    "name": "Budget 2026",
                    "mimeType": "application/vnd.google-apps.spreadsheet",
                    "webViewLink": "https://docs.google.com/spreadsheets/d/new1/edit",
                },
                None,
            )
        return {}, None

    async def fake_drive_post(path, params=None, json_body=None):
        assert path == "/files/tpl1/copy"
        assert json_body["name"] == "Budget 2026"
        return {"id": "new1", "name": "Budget 2026"}, None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    monkeypatch.setattr(sheets_server, "_drive_post", fake_drive_post)
    result = await sheets_server.create_spreadsheet_from_template("tpl1", "Budget 2026")
    assert "Spreadsheet created from template:" in result
    assert "New File ID: new1" in result


@pytest.mark.asyncio
async def test_import_csv_to_sheet(monkeypatch):
    async def fake_sheets_post(path, json_body=None, params=None):
        assert path.endswith(":clear")
        return {"clearedRange": "Sheet1!A:ZZZ"}, None

    async def fake_sheets_put(path, json_body=None, params=None):
        assert "/values/" in path
        assert json_body["values"][0] == ["name", "value"]
        return {
            "updatedRange": "Sheet1!A1:B2",
            "updatedRows": 2,
            "updatedColumns": 2,
            "updatedCells": 4,
        }, None

    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    monkeypatch.setattr(sheets_server, "_sheets_put", fake_sheets_put)
    result = await sheets_server.import_csv_to_sheet(
        "sp1",
        "Sheet1",
        "name,value\nops,42\n",
        overwrite=True,
    )
    assert "CSV import to sheet completed:" in result
    assert "Rows Imported: 2" in result


@pytest.mark.asyncio
async def test_insert_sheet_chart(monkeypatch):
    async def fake_sheets_post(path, json_body=None, params=None):
        assert path == "/sp1:batchUpdate"
        request = json_body["requests"][0]
        assert "addChart" in request
        return {"replies": [{"addChart": {"chart": {"chartId": 99}}}]}, None

    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    result = await sheets_server.insert_sheet_chart(
        "sp1",
        0,
        {"spec": {"title": "Revenue Chart", "basicChart": {"chartType": "COLUMN"}}},
    )
    assert "Sheet chart inserted:" in result
    assert "Chart ID: 99" in result


@pytest.mark.asyncio
async def test_protect_sheet_or_range_sheet_level(monkeypatch):
    async def fake_sheets_get(path, params=None):
        return (
            {
                "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp1/edit",
                "sheets": [{"properties": {"sheetId": 0, "title": "Sheet1"}}],
            },
            None,
        )

    async def fake_sheets_post(path, json_body=None, params=None):
        assert path == "/sp1:batchUpdate"
        req = json_body["requests"][0]["addProtectedRange"]["protectedRange"]
        assert req["range"]["sheetId"] == 0
        return {"replies": [{"addProtectedRange": {"protectedRange": {"protectedRangeId": 123}}}]}, None

    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    result = await sheets_server.protect_sheet_or_range(
        "sp1",
        sheet_id=0,
        editors=["alice@example.com"],
    )
    assert "Sheet protection created:" in result
    assert "Protected Range ID: 123" in result


@pytest.mark.asyncio
async def test_create_pivot_table(monkeypatch):
    async def fake_sheets_get(path, params=None):
        return (
            {
                "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/sp1/edit",
                "sheets": [
                    {"properties": {"sheetId": 0, "title": "Data"}},
                    {"properties": {"sheetId": 1, "title": "Summary"}},
                ],
            },
            None,
        )

    async def fake_sheets_post(path, json_body=None, params=None):
        assert path == "/sp1:batchUpdate"
        req = json_body["requests"][0]["updateCells"]
        assert req["start"]["sheetId"] == 1
        assert "pivotTable" in req["rows"][0]["values"][0]
        return {"replies": [{}]}, None

    monkeypatch.setattr(sheets_server, "_sheets_get", fake_sheets_get)
    monkeypatch.setattr(sheets_server, "_sheets_post", fake_sheets_post)
    result = await sheets_server.create_pivot_table(
        "sp1",
        source_range="Data!A1:B100",
        target_sheet="Summary",
        target_cell="A1",
    )
    assert "Pivot table created:" in result
    assert "Target Sheet: Summary" in result


@pytest.mark.asyncio
async def test_get_spreadsheet_permissions(monkeypatch):
    async def fake_drive_get(path, params=None):
        if path == "/files/sp1":
            return (
                {
                    "id": "sp1",
                    "name": "Budget",
                    "mimeType": "application/vnd.google-apps.spreadsheet",
                    "webViewLink": "https://docs.google.com/spreadsheets/d/sp1/edit",
                },
                None,
            )
        if path == "/files/sp1/permissions":
            return (
                {
                    "permissions": [
                        {
                            "id": "perm-1",
                            "type": "user",
                            "role": "writer",
                            "emailAddress": "alice@example.com",
                        }
                    ]
                },
                None,
            )
        return {}, None

    monkeypatch.setattr(sheets_server, "_drive_get", fake_drive_get)
    result = await sheets_server.get_spreadsheet_permissions("sp1")
    assert "Spreadsheet permissions:" in result
    assert "Permission Count: 1" in result
    assert "alice@example.com" in result
