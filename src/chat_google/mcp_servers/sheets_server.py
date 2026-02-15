import os
import json
import csv
from typing import Any
from io import BytesIO, StringIO
from datetime import datetime, timedelta, timezone
from urllib.parse import quote

import httpx
from dotenv import load_dotenv
from mcp.server.fastmcp import FastMCP
from pydantic import BaseModel, ConfigDict, Field, model_validator

load_dotenv()
mcp = FastMCP("GoogleSheets")

SHEETS_API_BASE = "https://sheets.googleapis.com/v4/spreadsheets"
DRIVE_API_BASE = "https://www.googleapis.com/drive/v3"
DRIVE_UPLOAD_API_BASE = "https://www.googleapis.com/upload/drive/v3"
GOOGLE_OAUTH_TOKEN_ENDPOINT = "https://oauth2.googleapis.com/token"
HTTP_TIMEOUT = httpx.Timeout(timeout=20.0, connect=5.0)
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXCEL_LEGACY_MIME = "application/vnd.ms-excel"
EXCEL_MIMES = {EXCEL_MIME, EXCEL_LEGACY_MIME}
TOKEN_EXPIRY_SAFETY_MARGIN_SECONDS = 60
VALUE_INPUT_OPTIONS = {"RAW", "USER_ENTERED"}
VALUE_RENDER_OPTIONS = {"FORMATTED_VALUE", "UNFORMATTED_VALUE", "FORMULA"}
DATE_TIME_RENDER_OPTIONS = {"SERIAL_NUMBER", "FORMATTED_STRING"}
SHEET_SHARE_ROLES = {"reader", "commenter", "writer"}
SHEET_CHART_TYPES = {"COLUMN", "BAR", "LINE", "AREA", "PIE", "SCATTER"}
PIVOT_SUMMARIZE_FUNCTIONS = {"SUM", "COUNTA", "COUNT", "AVERAGE", "MAX", "MIN"}
SHEET_EXPORT_FORMATS = {
    "xlsx": {
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "binary": True,
    },
    "ods": {"mime_type": "application/x-vnd.oasis.opendocument.spreadsheet", "binary": True},
    "pdf": {"mime_type": "application/pdf", "binary": True},
    "zip": {"mime_type": "application/zip", "binary": True},
    "csv": {"mime_type": "text/csv", "binary": False},
    "tsv": {"mime_type": "text/tab-separated-values", "binary": False},
}
DOCS_SHEETS_EXPORT_BASE = "https://docs.google.com/spreadsheets/d"

_CACHED_ACCESS_TOKEN: str | None = None
_CACHED_ACCESS_TOKEN_EXPIRES_AT: datetime | None = None


class _ListSheetsInput(BaseModel):
    limit: int = Field(default=10, ge=1, le=100, strict=True)


class _ListSpreadsheetsInput(BaseModel):
    limit: int = Field(default=10, ge=1, le=100, strict=True)
    include_excel: bool = True


class _SearchSheetsInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    query: str = Field(min_length=1)
    limit: int = Field(default=10, ge=1, le=100, strict=True)


class _SearchSpreadsheetsInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    query: str = Field(min_length=1)
    limit: int = Field(default=10, ge=1, le=100, strict=True)
    include_excel: bool = True


class _SpreadsheetIdInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    spreadsheet_id: str = Field(min_length=1)


class _ReadSheetValuesInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    spreadsheet_id: str = Field(min_length=1)
    range_a1: str = Field(default="A1:Z50", min_length=1)
    max_rows: int = Field(default=50, ge=1, le=500, strict=True)
    max_cols: int = Field(default=20, ge=1, le=200, strict=True)


class _AppendSheetRowInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=False)

    spreadsheet_id: str = Field(min_length=1)
    range_a1: str = Field(min_length=1)
    values: list[str] = Field(min_length=1, max_length=200)


class _UpdateSheetValuesInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=False)

    spreadsheet_id: str = Field(min_length=1)
    range_a1: str = Field(min_length=1)
    values: list[list[str]] = Field(min_length=1, max_length=1000)
    value_input_option: str = Field(default="USER_ENTERED")

    @model_validator(mode="after")
    def validate_rows(self):
        non_empty_row_found = any(isinstance(row, list) and len(row) > 0 for row in self.values)
        if not non_empty_row_found:
            raise ValueError("values must contain at least one non-empty row.")
        option = self.value_input_option.upper().strip()
        if option not in VALUE_INPUT_OPTIONS:
            raise ValueError(
                f"Invalid value_input_option '{self.value_input_option}'. "
                f"Allowed: {', '.join(sorted(VALUE_INPUT_OPTIONS))}"
            )
        self.value_input_option = option
        return self


class _CreateSpreadsheetInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    title: str = Field(min_length=1)
    sheet_title: str = Field(default="Sheet1", min_length=1)


class _AddSheetTabInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    spreadsheet_id: str = Field(min_length=1)
    title: str = Field(min_length=1)
    row_count: int = Field(default=1000, ge=1, le=2000000, strict=True)
    column_count: int = Field(default=26, ge=1, le=18278, strict=True)


class _SpreadsheetFileIdInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    file_id: str = Field(min_length=1)


class _ReadExcelValuesInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    file_id: str = Field(min_length=1)
    sheet_name: str = Field(default="")
    range_a1: str = Field(default="A1:Z50", min_length=1)
    max_rows: int = Field(default=50, ge=1, le=500, strict=True)
    max_cols: int = Field(default=20, ge=1, le=200, strict=True)


class _ConvertExcelInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    file_id: str = Field(min_length=1)
    new_title: str = Field(default="")
    move_to_parent: bool = True


class _ExportGoogleSheetInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    file_id: str = Field(min_length=1)
    export_format: str = Field(default="xlsx")
    gid: str | None = None
    range_a1: str = Field(default="")
    max_preview_chars: int = Field(default=2000, ge=200, le=12000, strict=True)

    @model_validator(mode="after")
    def validate_values(self):
        normalized_format = self.export_format.lower().strip()
        if normalized_format not in SHEET_EXPORT_FORMATS:
            raise ValueError(
                f"Unsupported export_format '{self.export_format}'. "
                f"Allowed: {', '.join(sorted(SHEET_EXPORT_FORMATS.keys()))}"
            )
        self.export_format = normalized_format
        self.gid = self.gid.strip() if isinstance(self.gid, str) else None
        self.range_a1 = self.range_a1.strip()
        return self


class _BatchGetSheetValuesInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    spreadsheet_id: str = Field(min_length=1)
    ranges: list[str] = Field(min_length=1, max_length=50)
    value_render_option: str = Field(default="FORMATTED_VALUE")
    date_time_render_option: str = Field(default="SERIAL_NUMBER")
    max_rows_per_range: int = Field(default=20, ge=1, le=200, strict=True)
    max_cols_per_row: int = Field(default=20, ge=1, le=200, strict=True)

    @model_validator(mode="after")
    def validate_values(self):
        cleaned_ranges: list[str] = []
        for raw in self.ranges:
            cleaned = str(raw).strip()
            if cleaned:
                cleaned_ranges.append(cleaned)
        if not cleaned_ranges:
            raise ValueError("ranges must include at least one non-empty A1 range.")
        self.ranges = cleaned_ranges

        normalized_value_render = self.value_render_option.upper().strip()
        if normalized_value_render not in VALUE_RENDER_OPTIONS:
            raise ValueError(
                f"Invalid value_render_option '{self.value_render_option}'. "
                f"Allowed: {', '.join(sorted(VALUE_RENDER_OPTIONS))}"
            )
        self.value_render_option = normalized_value_render

        normalized_date_time_render = self.date_time_render_option.upper().strip()
        if normalized_date_time_render not in DATE_TIME_RENDER_OPTIONS:
            raise ValueError(
                f"Invalid date_time_render_option '{self.date_time_render_option}'. "
                f"Allowed: {', '.join(sorted(DATE_TIME_RENDER_OPTIONS))}"
            )
        self.date_time_render_option = normalized_date_time_render
        return self


class _BatchUpdateRangeInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=False)

    range_a1: str = Field(min_length=1)
    values: list[list[Any]] = Field(min_length=1, max_length=2000)

    @model_validator(mode="after")
    def validate_values(self):
        cleaned_range = self.range_a1.strip()
        if not cleaned_range:
            raise ValueError("range_a1 must not be empty.")
        self.range_a1 = cleaned_range

        non_empty_row_found = any(isinstance(row, list) and len(row) > 0 for row in self.values)
        if not non_empty_row_found:
            raise ValueError("values must contain at least one non-empty row.")
        return self


class _BatchUpdateSheetValuesInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=False)

    spreadsheet_id: str = Field(min_length=1)
    updates: list[_BatchUpdateRangeInput] = Field(min_length=1, max_length=100)
    value_input_option: str = Field(default="USER_ENTERED")

    @model_validator(mode="after")
    def validate_values(self):
        normalized_option = self.value_input_option.upper().strip()
        if normalized_option not in VALUE_INPUT_OPTIONS:
            raise ValueError(
                f"Invalid value_input_option '{self.value_input_option}'. "
                f"Allowed: {', '.join(sorted(VALUE_INPUT_OPTIONS))}"
            )
        self.value_input_option = normalized_option
        return self


class _ShareSpreadsheetInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    file_id: str = Field(min_length=1)
    user_email: str = Field(min_length=5)
    role: str = Field(default="writer")
    send_notification: bool = True
    message: str = Field(default="")

    @model_validator(mode="after")
    def validate_values(self):
        email = self.user_email.strip()
        if "@" not in email or email.startswith("@") or email.endswith("@"):
            raise ValueError("user_email must be a valid email address.")
        self.user_email = email

        normalized_role = self.role.lower().strip()
        if normalized_role not in SHEET_SHARE_ROLES:
            raise ValueError(
                f"Invalid role '{self.role}'. Allowed: {', '.join(sorted(SHEET_SHARE_ROLES))}"
            )
        self.role = normalized_role
        return self


class _CreateSpreadsheetFromTemplateInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    template_file_id: str = Field(min_length=1)
    new_title: str = Field(min_length=1)
    destination_folder_id: str = Field(default="")


class _ImportCsvToSheetInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=False)

    spreadsheet_id: str = Field(min_length=1)
    sheet_name: str = Field(min_length=1)
    csv_text: str = Field(min_length=1)
    overwrite: bool = False


class _InsertSheetChartInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    spreadsheet_id: str = Field(min_length=1)
    sheet_id: int = Field(ge=0, strict=True)
    chart_spec: dict[str, Any] = Field(default_factory=dict)

    @model_validator(mode="after")
    def validate_values(self):
        if not self.chart_spec:
            raise ValueError("chart_spec must not be empty.")
        return self


class _ProtectSheetOrRangeInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    spreadsheet_id: str = Field(min_length=1)
    sheet_id: int | None = Field(default=None, ge=0, strict=True)
    range_a1: str = Field(default="")
    editors: list[str] = Field(default_factory=list, max_length=100)
    warning_only: bool = False

    @model_validator(mode="after")
    def validate_values(self):
        self.range_a1 = self.range_a1.strip()
        if self.sheet_id is None and not self.range_a1:
            raise ValueError("Provide either sheet_id or range_a1.")
        cleaned_editors: list[str] = []
        for raw in self.editors:
            email = str(raw).strip()
            if not email:
                continue
            if "@" not in email or email.startswith("@") or email.endswith("@"):
                raise ValueError(f"Invalid editor email: '{raw}'")
            cleaned_editors.append(email)
        self.editors = cleaned_editors
        return self


class _CreatePivotTableInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    spreadsheet_id: str = Field(min_length=1)
    source_range: str = Field(min_length=1)
    target_sheet: str = Field(min_length=1)
    target_cell: str = Field(default="A1", min_length=1)
    summarize_function: str = Field(default="COUNTA")

    @model_validator(mode="after")
    def validate_values(self):
        normalized = self.summarize_function.upper().strip()
        if normalized not in PIVOT_SUMMARIZE_FUNCTIONS:
            raise ValueError(
                f"Invalid summarize_function '{self.summarize_function}'. "
                f"Allowed: {', '.join(sorted(PIVOT_SUMMARIZE_FUNCTIONS))}"
            )
        self.summarize_function = normalized
        return self


def _get_cached_access_token() -> str | None:
    global _CACHED_ACCESS_TOKEN, _CACHED_ACCESS_TOKEN_EXPIRES_AT
    if not _CACHED_ACCESS_TOKEN or not _CACHED_ACCESS_TOKEN_EXPIRES_AT:
        return None
    if datetime.now(timezone.utc) >= _CACHED_ACCESS_TOKEN_EXPIRES_AT:
        _CACHED_ACCESS_TOKEN = None
        _CACHED_ACCESS_TOKEN_EXPIRES_AT = None
        return None
    return _CACHED_ACCESS_TOKEN


def _invalidate_cached_access_token() -> None:
    global _CACHED_ACCESS_TOKEN, _CACHED_ACCESS_TOKEN_EXPIRES_AT
    _CACHED_ACCESS_TOKEN = None
    _CACHED_ACCESS_TOKEN_EXPIRES_AT = None


def _set_cached_access_token(token: str, expires_in_seconds: int) -> None:
    global _CACHED_ACCESS_TOKEN, _CACHED_ACCESS_TOKEN_EXPIRES_AT
    safe_ttl = max(0, int(expires_in_seconds) - TOKEN_EXPIRY_SAFETY_MARGIN_SECONDS)
    _CACHED_ACCESS_TOKEN = token
    _CACHED_ACCESS_TOKEN_EXPIRES_AT = datetime.now(timezone.utc) + timedelta(seconds=safe_ttl)


def _client_kwargs() -> dict:
    return {"follow_redirects": True, "timeout": HTTP_TIMEOUT}


def _refresh_access_token(
    refresh_token: str,
    client_id: str,
    client_secret: str,
) -> tuple[str, int]:
    payload = {
        "grant_type": "refresh_token",
        "refresh_token": refresh_token,
        "client_id": client_id,
        "client_secret": client_secret,
    }
    with httpx.Client(**_client_kwargs()) as client:
        response = client.post(GOOGLE_OAUTH_TOKEN_ENDPOINT, data=payload)

    if response.status_code != 200:
        detail = ""
        try:
            body = response.json()
            error = body.get("error", "")
            error_description = body.get("error_description", "")
            detail = f"{error}: {error_description}".strip(": ").strip()
        except Exception:
            detail = response.text.strip()[:300]
        detail_part = f" - {detail}" if detail else ""
        raise ValueError(
            f"Sheets OAuth refresh failed with HTTP {response.status_code}{detail_part}"
        )

    try:
        data = response.json()
    except Exception as exc:
        raise ValueError(f"Sheets OAuth refresh response parse error: {exc}") from exc

    token = str(data.get("access_token", "")).strip()
    if not token:
        raise ValueError("Sheets OAuth refresh response missing access_token")

    expires_in_raw = data.get("expires_in", 3600)
    try:
        expires_in = int(expires_in_raw)
    except Exception:
        expires_in = 3600
    return token, expires_in


def _get_access_token() -> str:
    load_dotenv(override=True)

    cached = _get_cached_access_token()
    if cached:
        return cached

    static_token = (os.getenv("GOOGLE_DRIVE_ACCESS_TOKEN") or "").strip()
    refresh_token = (os.getenv("GOOGLE_DRIVE_REFRESH_TOKEN") or "").strip()
    client_id = (os.getenv("GOOGLE_OAUTH_CLIENT_ID") or "").strip()
    client_secret = (os.getenv("GOOGLE_OAUTH_CLIENT_SECRET") or "").strip()

    has_any_refresh_inputs = any([refresh_token, client_id, client_secret])
    has_full_refresh_inputs = all([refresh_token, client_id, client_secret])

    if has_full_refresh_inputs:
        try:
            refreshed_token, expires_in = _refresh_access_token(
                refresh_token=refresh_token,
                client_id=client_id,
                client_secret=client_secret,
            )
            _set_cached_access_token(refreshed_token, expires_in)
            os.environ["GOOGLE_DRIVE_ACCESS_TOKEN"] = refreshed_token
            return refreshed_token
        except Exception as exc:
            if static_token:
                return static_token
            raise ValueError(f"Failed to refresh Sheets access token: {exc}") from exc

    if has_any_refresh_inputs and not has_full_refresh_inputs:
        missing = []
        if not refresh_token:
            missing.append("GOOGLE_DRIVE_REFRESH_TOKEN")
        if not client_id:
            missing.append("GOOGLE_OAUTH_CLIENT_ID")
        if not client_secret:
            missing.append("GOOGLE_OAUTH_CLIENT_SECRET")
        if static_token:
            return static_token
        raise ValueError(
            "Incomplete Sheets OAuth refresh configuration. Missing: " + ", ".join(missing)
        )

    if static_token:
        return static_token

    raise ValueError(
        "Set GOOGLE_DRIVE_ACCESS_TOKEN or configure refresh flow with "
        "GOOGLE_DRIVE_REFRESH_TOKEN, GOOGLE_OAUTH_CLIENT_ID, and GOOGLE_OAUTH_CLIENT_SECRET in .env"
    )


def _auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _escape_query(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def _encode_a1_range(range_a1: str) -> str:
    return quote(range_a1, safe="!:$',()")


def _format_sheets_error(response: httpx.Response) -> str:
    status = response.status_code
    detail = ""
    reason = ""
    try:
        payload = response.json()
        error_obj = payload.get("error", {}) if isinstance(payload, dict) else {}
        detail = str(error_obj.get("message", "")).strip()
        errors = error_obj.get("errors", []) if isinstance(error_obj, dict) else []
        if isinstance(errors, list) and errors:
            first = errors[0]
            if isinstance(first, dict):
                reason = str(first.get("reason", "")).strip()
    except Exception:
        detail = response.text.strip()[:300]

    hint = ""
    if status == 401:
        hint = (
            " Hint: access token expired/invalid. Configure refresh flow with "
            "GOOGLE_DRIVE_REFRESH_TOKEN, GOOGLE_OAUTH_CLIENT_ID, GOOGLE_OAUTH_CLIENT_SECRET."
        )
    elif status == 403:
        hint = (
            " Hint: ensure Google Sheets API is enabled and token scope allows Sheets/Drive access."
        )
    reason_part = f" ({reason})" if reason else ""
    detail_part = f" - {detail}" if detail else ""
    return f"Error: Google Sheets API request failed: {status}{reason_part}{detail_part}.{hint}".strip()


def _format_drive_error(response: httpx.Response) -> str:
    status = response.status_code
    detail = ""
    reason = ""
    try:
        payload = response.json()
        error_obj = payload.get("error", {}) if isinstance(payload, dict) else {}
        detail = str(error_obj.get("message", "")).strip()
        errors = error_obj.get("errors", []) if isinstance(error_obj, dict) else []
        if isinstance(errors, list) and errors:
            first = errors[0]
            if isinstance(first, dict):
                reason = str(first.get("reason", "")).strip()
    except Exception:
        detail = response.text.strip()[:300]

    hint = ""
    if status == 401:
        hint = (
            " Hint: access token expired/invalid. Configure refresh flow with "
            "GOOGLE_DRIVE_REFRESH_TOKEN, GOOGLE_OAUTH_CLIENT_ID, GOOGLE_OAUTH_CLIENT_SECRET."
        )
    reason_part = f" ({reason})" if reason else ""
    detail_part = f" - {detail}" if detail else ""
    return f"Error: Drive API request failed: {status}{reason_part}{detail_part}.{hint}".strip()


def _format_sheet_file_line(item: dict) -> str:
    name = item.get("name", "Untitled")
    sheet_id = item.get("id", "-")
    modified = item.get("modifiedTime", "-")
    link = item.get("webViewLink", "-")
    return f"- {name} | ID: {sheet_id} | Modified: {modified} | Link: {link}"


def _spreadsheet_query(include_excel: bool) -> str:
    if include_excel:
        return (
            "("
            f"mimeType='{GOOGLE_SHEET_MIME}' or "
            f"mimeType='{EXCEL_MIME}' or "
            f"mimeType='{EXCEL_LEGACY_MIME}'"
            ") and trashed=false"
        )
    return f"mimeType='{GOOGLE_SHEET_MIME}' and trashed=false"


def _mime_label(mime_type: str) -> str:
    if mime_type == GOOGLE_SHEET_MIME:
        return "google_sheet"
    if mime_type == EXCEL_MIME:
        return "excel_xlsx"
    if mime_type == EXCEL_LEGACY_MIME:
        return "excel_xls"
    return mime_type or "-"


def _format_spreadsheet_file_line(item: dict) -> str:
    name = item.get("name", "Untitled")
    file_id = item.get("id", "-")
    modified = item.get("modifiedTime", "-")
    link = item.get("webViewLink", "-")
    mime_type = str(item.get("mimeType", "")).strip()
    return (
        f"- {name} | ID: {file_id} | Type: {_mime_label(mime_type)} "
        f"| Modified: {modified} | Link: {link}"
    )


async def _sheets_get(path: str, params: dict | None = None) -> tuple[dict | None, str | None]:
    token = _get_access_token()
    url = f"{SHEETS_API_BASE}{path}"
    async with httpx.AsyncClient(**_client_kwargs()) as client:
        response = await client.get(url, headers=_auth_headers(token), params=params)
        if response.status_code == 401:
            _invalidate_cached_access_token()
            retry_token = _get_access_token()
            if retry_token:
                response = await client.get(url, headers=_auth_headers(retry_token), params=params)
    if response.status_code != 200:
        return None, _format_sheets_error(response)
    try:
        return response.json(), None
    except Exception as exc:
        return None, f"Google Sheets API response parse error: {str(exc)}"


async def _sheets_post(
    path: str,
    json_body: dict | None = None,
    params: dict | None = None,
) -> tuple[dict | None, str | None]:
    token = _get_access_token()
    url = f"{SHEETS_API_BASE}{path}"
    async with httpx.AsyncClient(**_client_kwargs()) as client:
        response = await client.post(url, headers=_auth_headers(token), json=json_body, params=params)
        if response.status_code == 401:
            _invalidate_cached_access_token()
            retry_token = _get_access_token()
            if retry_token:
                response = await client.post(
                    url,
                    headers=_auth_headers(retry_token),
                    json=json_body,
                    params=params,
                )
    if response.status_code not in (200, 201):
        return None, _format_sheets_error(response)
    try:
        return response.json(), None
    except Exception:
        return {}, None


async def _sheets_put(
    path: str,
    json_body: dict | None = None,
    params: dict | None = None,
) -> tuple[dict | None, str | None]:
    token = _get_access_token()
    url = f"{SHEETS_API_BASE}{path}"
    async with httpx.AsyncClient(**_client_kwargs()) as client:
        response = await client.put(url, headers=_auth_headers(token), json=json_body, params=params)
        if response.status_code == 401:
            _invalidate_cached_access_token()
            retry_token = _get_access_token()
            if retry_token:
                response = await client.put(
                    url,
                    headers=_auth_headers(retry_token),
                    json=json_body,
                    params=params,
                )
    if response.status_code != 200:
        return None, _format_sheets_error(response)
    try:
        return response.json(), None
    except Exception:
        return {}, None


async def _drive_get(path: str, params: dict | None = None) -> tuple[dict | None, str | None]:
    token = _get_access_token()
    url = f"{DRIVE_API_BASE}{path}"
    async with httpx.AsyncClient(**_client_kwargs()) as client:
        response = await client.get(url, headers=_auth_headers(token), params=params)
        if response.status_code == 401:
            _invalidate_cached_access_token()
            retry_token = _get_access_token()
            if retry_token:
                response = await client.get(url, headers=_auth_headers(retry_token), params=params)
    if response.status_code != 200:
        return None, _format_drive_error(response)
    try:
        return response.json(), None
    except Exception as exc:
        return None, f"Drive API response parse error: {str(exc)}"


async def _drive_post(
    path: str,
    params: dict | None = None,
    json_body: dict | None = None,
) -> tuple[dict | None, str | None]:
    token = _get_access_token()
    url = f"{DRIVE_API_BASE}{path}"
    async with httpx.AsyncClient(**_client_kwargs()) as client:
        response = await client.post(url, headers=_auth_headers(token), params=params, json=json_body)
        if response.status_code == 401:
            _invalidate_cached_access_token()
            retry_token = _get_access_token()
            if retry_token:
                response = await client.post(
                    url,
                    headers=_auth_headers(retry_token),
                    params=params,
                    json=json_body,
                )
    if response.status_code not in (200, 201):
        return None, _format_drive_error(response)
    try:
        return response.json(), None
    except Exception:
        return {}, None


async def _drive_get_bytes(path: str, params: dict | None = None) -> tuple[bytes | None, str | None]:
    token = _get_access_token()
    url = f"{DRIVE_API_BASE}{path}"
    async with httpx.AsyncClient(**_client_kwargs()) as client:
        response = await client.get(url, headers=_auth_headers(token), params=params)
        if response.status_code == 401:
            _invalidate_cached_access_token()
            retry_token = _get_access_token()
            if retry_token:
                response = await client.get(url, headers=_auth_headers(retry_token), params=params)
    if response.status_code != 200:
        return None, _format_drive_error(response)
    return response.content, None


async def _drive_upload_multipart(
    metadata: dict,
    media_bytes: bytes,
    media_mime: str,
    media_filename: str = "source.bin",
) -> tuple[dict | None, str | None]:
    token = _get_access_token()
    url = f"{DRIVE_UPLOAD_API_BASE}/files"
    params = {"uploadType": "multipart", "supportsAllDrives": "true"}
    files = {
        "metadata": (
            "metadata",
            json.dumps(metadata),
            "application/json; charset=UTF-8",
        ),
        "media": (media_filename, media_bytes, media_mime),
    }
    async with httpx.AsyncClient(**_client_kwargs()) as client:
        response = await client.post(url, headers=_auth_headers(token), params=params, files=files)
        if response.status_code == 401:
            _invalidate_cached_access_token()
            retry_token = _get_access_token()
            if retry_token:
                response = await client.post(
                    url,
                    headers=_auth_headers(retry_token),
                    params=params,
                    files=files,
                )
    if response.status_code not in (200, 201):
        return None, _format_drive_error(response)
    try:
        return response.json(), None
    except Exception:
        return {}, None


async def _drive_export_google_sheet_bytes(
    file_id: str,
    export_format: str,
    mime_type: str,
    gid: str | None = None,
    range_a1: str = "",
) -> tuple[bytes | None, str | None]:
    if export_format in {"csv", "tsv"} and gid:
        token = _get_access_token()
        params: dict[str, str] = {"format": export_format, "gid": gid}
        if range_a1:
            params["range"] = range_a1
        url = f"{DOCS_SHEETS_EXPORT_BASE}/{file_id}/export"
        async with httpx.AsyncClient(**_client_kwargs()) as client:
            response = await client.get(url, headers=_auth_headers(token), params=params)
            if response.status_code == 401:
                _invalidate_cached_access_token()
                retry_token = _get_access_token()
                if retry_token:
                    response = await client.get(
                        url,
                        headers=_auth_headers(retry_token),
                        params=params,
                    )
        if response.status_code != 200:
            return None, _format_drive_error(response)
        return response.content, None

    return await _drive_get_bytes(
        f"/files/{file_id}/export",
        params={"mimeType": mime_type},
    )


def _decode_text_with_fallback(payload: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return payload.decode(encoding)
        except Exception:
            continue
    return payload.decode("utf-8", errors="replace")


def _quote_sheet_name(sheet_name: str) -> str:
    cleaned = sheet_name.strip()
    escaped = cleaned.replace("'", "''")
    return f"'{escaped}'"


async def _get_sheet_mappings(
    spreadsheet_id: str,
) -> tuple[dict[str, int], dict[int, str], str, str | None]:
    data, err = await _sheets_get(
        f"/{spreadsheet_id}",
        params={"fields": "spreadsheetUrl,sheets(properties.sheetId,properties.title)"},
    )
    if err:
        return {}, {}, "-", err
    if not data:
        return {}, {}, "-", "Spreadsheet metadata not found."

    title_to_id: dict[str, int] = {}
    id_to_title: dict[int, str] = {}
    sheets = data.get("sheets", []) if isinstance(data, dict) else []
    for item in sheets:
        props = item.get("properties", {}) if isinstance(item, dict) else {}
        title = str(props.get("title", "")).strip()
        sheet_id = props.get("sheetId")
        if title and isinstance(sheet_id, int):
            title_to_id[title.lower()] = sheet_id
            id_to_title[sheet_id] = title

    return title_to_id, id_to_title, str(data.get("spreadsheetUrl", "-")), None


def _grid_range_from_body(sheet_id: int, body_range: str) -> tuple[dict | None, str | None]:
    try:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(body_range)
    except Exception:
        return None, (
            f"Unable to parse range '{body_range}'. Use standard A1 notation like A1:C20."
        )

    return (
        {
            "sheetId": sheet_id,
            "startRowIndex": min_row - 1,
            "endRowIndex": max_row,
            "startColumnIndex": min_col - 1,
            "endColumnIndex": max_col,
        },
        None,
    )


def _a1_cell_to_index(cell_a1: str) -> tuple[int | None, int | None, str | None]:
    raw = cell_a1.strip()
    if not raw:
        return None, None, "target_cell must not be empty."
    try:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, _, _ = range_boundaries(raw)
    except Exception:
        return None, None, (
            f"Unable to parse target_cell '{cell_a1}'. Use single-cell A1 notation like A1."
        )
    return min_row - 1, min_col - 1, None


async def _resolve_grid_range_from_a1(
    spreadsheet_id: str,
    range_a1: str,
    fallback_sheet_id: int | None = None,
) -> tuple[dict | None, str | None, str | None]:
    sheet_name, body_range = _split_sheet_range(range_a1)
    title_to_id, id_to_title, _, map_err = await _get_sheet_mappings(spreadsheet_id)
    if map_err:
        return None, None, map_err

    sheet_id = fallback_sheet_id
    resolved_name: str | None = None
    if sheet_name:
        resolved = title_to_id.get(sheet_name.lower())
        if resolved is None:
            return None, None, f"Sheet '{sheet_name}' not found in spreadsheet."
        sheet_id = resolved
        resolved_name = id_to_title.get(resolved, sheet_name)
    elif sheet_id is not None:
        resolved_name = id_to_title.get(sheet_id)
    else:
        return None, None, "Sheet name is required in range_a1 when sheet_id is not provided."

    if sheet_id is None:
        return None, None, "Unable to resolve target sheet."

    grid_range, parse_err = _grid_range_from_body(sheet_id, body_range)
    if parse_err:
        return None, None, parse_err
    return grid_range, resolved_name, None


def _parse_excel_bounds(range_a1: str) -> tuple[int, int, int, int]:
    try:
        from openpyxl.utils.cell import range_boundaries

        return range_boundaries(range_a1)
    except Exception:
        # fallback to first 26 columns and first 50 rows
        return 1, 1, 26, 50


def _split_sheet_range(range_a1: str) -> tuple[str | None, str]:
    raw = range_a1.strip()
    if "!" not in raw:
        return None, raw
    sheet, tail = raw.split("!", 1)
    cleaned_sheet = sheet.strip().strip("'")
    return (cleaned_sheet or None), tail.strip() or "A1:Z50"


def _excel_cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


@mcp.tool()
async def list_sheets_spreadsheets(limit: int = 10) -> str:
    """Lists Google Sheets files from Drive."""
    try:
        params = _ListSheetsInput.model_validate({"limit": limit})
        data, err = await _drive_get(
            "/files",
            params={
                "q": f"mimeType='{GOOGLE_SHEET_MIME}' and trashed=false",
                "orderBy": "modifiedTime desc",
                "pageSize": params.limit,
                "fields": "files(id,name,modifiedTime,webViewLink),nextPageToken",
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
            },
        )
        if err:
            return err
        files = data.get("files", []) if data else []
        if not files:
            return "No Google Sheets files found."
        lines = [_format_sheet_file_line(item) for item in files]
        return f"Google Sheets files (showing {len(lines)}):\n" + "\n".join(lines)
    except Exception as exc:
        return f"Error listing Google Sheets files: {str(exc)}"


@mcp.tool()
async def search_sheets_spreadsheets(query: str, limit: int = 10) -> str:
    """Searches Google Sheets files by title/name."""
    try:
        params = _SearchSheetsInput.model_validate({"query": query, "limit": limit})
        safe_query = _escape_query(params.query)
        q = (
            f"mimeType='{GOOGLE_SHEET_MIME}' and trashed=false and "
            f"name contains '{safe_query}'"
        )
        data, err = await _drive_get(
            "/files",
            params={
                "q": q,
                "orderBy": "modifiedTime desc",
                "pageSize": params.limit,
                "fields": "files(id,name,modifiedTime,webViewLink),nextPageToken",
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
            },
        )
        if err:
            return err
        files = data.get("files", []) if data else []
        if not files:
            return f"No Google Sheets files found matching '{params.query}'"
        lines = [_format_sheet_file_line(item) for item in files]
        return (
            f"Google Sheets search results for '{params.query}' (showing {len(lines)}):\n"
            + "\n".join(lines)
        )
    except Exception as exc:
        return f"Error searching Google Sheets files: {str(exc)}"


@mcp.tool()
async def list_spreadsheets(limit: int = 10, include_excel: bool = True) -> str:
    """Lists spreadsheets, optionally including .xlsx/.xls files."""
    try:
        params = _ListSpreadsheetsInput.model_validate(
            {"limit": limit, "include_excel": include_excel}
        )
        data, err = await _drive_get(
            "/files",
            params={
                "q": _spreadsheet_query(params.include_excel),
                "orderBy": "modifiedTime desc",
                "pageSize": params.limit,
                "fields": "files(id,name,mimeType,modifiedTime,webViewLink),nextPageToken",
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
            },
        )
        if err:
            return err
        files = data.get("files", []) if data else []
        if not files:
            return "No spreadsheet files found."
        lines = [_format_spreadsheet_file_line(item) for item in files]
        return (
            f"Spreadsheet files (include_excel={params.include_excel}, showing {len(lines)}):\n"
            + "\n".join(lines)
        )
    except Exception as exc:
        return f"Error listing spreadsheet files: {str(exc)}"


@mcp.tool()
async def search_spreadsheets(query: str, limit: int = 10, include_excel: bool = True) -> str:
    """Searches spreadsheets by title, optionally including .xlsx/.xls files."""
    try:
        params = _SearchSpreadsheetsInput.model_validate(
            {"query": query, "limit": limit, "include_excel": include_excel}
        )
        safe_query = _escape_query(params.query)
        q = _spreadsheet_query(params.include_excel) + f" and name contains '{safe_query}'"
        data, err = await _drive_get(
            "/files",
            params={
                "q": q,
                "orderBy": "modifiedTime desc",
                "pageSize": params.limit,
                "fields": "files(id,name,mimeType,modifiedTime,webViewLink),nextPageToken",
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
            },
        )
        if err:
            return err
        files = data.get("files", []) if data else []
        if not files:
            return f"No spreadsheet files found matching '{params.query}'"
        lines = [_format_spreadsheet_file_line(item) for item in files]
        return (
            f"Spreadsheet search results for '{params.query}' "
            f"(include_excel={params.include_excel}, showing {len(lines)}):\n"
            + "\n".join(lines)
        )
    except Exception as exc:
        return f"Error searching spreadsheet files: {str(exc)}"


@mcp.tool()
async def get_sheets_metadata(spreadsheet_id: str) -> str:
    """Gets spreadsheet metadata including sheet/tab list."""
    try:
        params = _SpreadsheetIdInput.model_validate({"spreadsheet_id": spreadsheet_id})
        data, err = await _sheets_get(
            f"/{params.spreadsheet_id}",
            params={
                "fields": (
                    "spreadsheetId,spreadsheetUrl,properties.title,"
                    "sheets(properties.sheetId,properties.title,properties.gridProperties)"
                )
            },
        )
        if err:
            return err
        if not data:
            return "Spreadsheet metadata not found."

        sheets = data.get("sheets", []) if isinstance(data, dict) else []
        lines = []
        for item in sheets:
            props = item.get("properties", {}) if isinstance(item, dict) else {}
            grid = props.get("gridProperties", {}) if isinstance(props, dict) else {}
            lines.append(
                f"- {props.get('title', '-')}"
                f" | Sheet ID: {props.get('sheetId', '-')}"
                f" | Grid: {grid.get('rowCount', '-')}x{grid.get('columnCount', '-')}"
            )
        sheet_lines = "\n".join(lines) if lines else "- (none)"

        return (
            "Google Sheets Metadata:\n"
            f"Title: {(data.get('properties', {}) or {}).get('title', '-')}\n"
            f"Spreadsheet ID: {data.get('spreadsheetId', params.spreadsheet_id)}\n"
            f"Spreadsheet URL: {data.get('spreadsheetUrl', '-')}\n"
            f"Total Tabs: {len(sheets)}\n"
            f"Tabs:\n{sheet_lines}"
        )
    except Exception as exc:
        return f"Error getting Google Sheets metadata: {str(exc)}"


@mcp.tool()
async def read_sheet_values(
    spreadsheet_id: str,
    range_a1: str = "A1:Z50",
    max_rows: int = 50,
    max_cols: int = 20,
) -> str:
    """Reads cell values from a spreadsheet range."""
    try:
        params = _ReadSheetValuesInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "range_a1": range_a1,
                "max_rows": max_rows,
                "max_cols": max_cols,
            }
        )
        encoded_range = _encode_a1_range(params.range_a1)
        data, err = await _sheets_get(f"/{params.spreadsheet_id}/values/{encoded_range}")
        if err:
            return err
        if not data:
            return "No value response from Google Sheets."

        values = data.get("values", []) if isinstance(data, dict) else []
        if not values:
            return (
                f"No values found for spreadsheet '{params.spreadsheet_id}' "
                f"range '{params.range_a1}'."
            )

        total_rows = len(values)
        displayed_rows = values[: params.max_rows]
        lines = []
        cols_truncated = False
        for idx, row in enumerate(displayed_rows, start=1):
            row_values = [str(v) for v in row[: params.max_cols]]
            if len(row) > params.max_cols:
                cols_truncated = True
            lines.append(f"- R{idx}: {' | '.join(row_values)}")

        notes = []
        if total_rows > params.max_rows:
            notes.append(f"Row display limited to first {params.max_rows} rows.")
        if cols_truncated:
            notes.append(f"Column display limited to first {params.max_cols} columns per row.")
        notes_text = ("\nNotes:\n- " + "\n- ".join(notes)) if notes else ""

        return (
            "Google Sheets Values:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Requested Range: {params.range_a1}\n"
            f"Returned Range: {data.get('range', params.range_a1)}\n"
            f"Rows Returned: {total_rows}\n"
            f"Rows Shown: {len(displayed_rows)}\n"
            f"Values:\n{chr(10).join(lines)}{notes_text}"
        )
    except Exception as exc:
        return f"Error reading Google Sheets values: {str(exc)}"


@mcp.tool()
async def append_sheet_row(
    spreadsheet_id: str,
    range_a1: str,
    values: list[str],
) -> str:
    """Appends a single row to a sheet range."""
    try:
        params = _AppendSheetRowInput.model_validate(
            {"spreadsheet_id": spreadsheet_id, "range_a1": range_a1, "values": values}
        )
        encoded_range = _encode_a1_range(params.range_a1)
        data, err = await _sheets_post(
            f"/{params.spreadsheet_id}/values/{encoded_range}:append",
            params={
                "valueInputOption": "USER_ENTERED",
                "insertDataOption": "INSERT_ROWS",
            },
            json_body={"majorDimension": "ROWS", "values": [[str(v) for v in params.values]]},
        )
        if err:
            return err

        updates = data.get("updates", {}) if isinstance(data, dict) else {}
        meta, meta_err = await _sheets_get(
            f"/{params.spreadsheet_id}",
            params={"fields": "spreadsheetUrl,properties.title"},
        )
        if meta_err:
            meta = {}
        return (
            "Google Sheets append row completed:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Spreadsheet Title: {(meta.get('properties', {}) or {}).get('title', '-')}\n"
            f"Updated Range: {updates.get('updatedRange', '-')}\n"
            f"Updated Rows: {updates.get('updatedRows', '-')}\n"
            f"Updated Cells: {updates.get('updatedCells', '-')}\n"
            f"Spreadsheet URL: {meta.get('spreadsheetUrl', '-')}"
        )
    except Exception as exc:
        return f"Error appending Google Sheets row: {str(exc)}"


@mcp.tool()
async def update_sheet_values(
    spreadsheet_id: str,
    range_a1: str,
    values: list[list[str]],
    value_input_option: str = "USER_ENTERED",
) -> str:
    """Updates a fixed range with provided 2D values."""
    try:
        params = _UpdateSheetValuesInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "range_a1": range_a1,
                "values": values,
                "value_input_option": value_input_option,
            }
        )
        encoded_range = _encode_a1_range(params.range_a1)
        normalized_values = [[str(cell) for cell in row] for row in params.values]
        data, err = await _sheets_put(
            f"/{params.spreadsheet_id}/values/{encoded_range}",
            params={"valueInputOption": params.value_input_option},
            json_body={"majorDimension": "ROWS", "values": normalized_values},
        )
        if err:
            return err
        return (
            "Google Sheets update values completed:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Updated Range: {data.get('updatedRange', '-') if data else '-'}\n"
            f"Updated Rows: {data.get('updatedRows', '-') if data else '-'}\n"
            f"Updated Columns: {data.get('updatedColumns', '-') if data else '-'}\n"
            f"Updated Cells: {data.get('updatedCells', '-') if data else '-'}\n"
            f"Value Input Option: {params.value_input_option}"
        )
    except Exception as exc:
        return f"Error updating Google Sheets values: {str(exc)}"


@mcp.tool()
async def create_sheets_spreadsheet(title: str, sheet_title: str = "Sheet1") -> str:
    """Creates a new spreadsheet with an initial tab."""
    try:
        params = _CreateSpreadsheetInput.model_validate({"title": title, "sheet_title": sheet_title})
        data, err = await _sheets_post(
            "",
            json_body={
                "properties": {"title": params.title},
                "sheets": [{"properties": {"title": params.sheet_title}}],
            },
        )
        if err:
            return err
        if not data:
            return "Failed to create Google Sheets spreadsheet."

        spreadsheet_id = data.get("spreadsheetId", "-")
        spreadsheet_url = data.get("spreadsheetUrl", "-")
        sheets = data.get("sheets", []) if isinstance(data, dict) else []
        first_sheet_props = (
            (sheets[0].get("properties", {}) if sheets and isinstance(sheets[0], dict) else {})
            if isinstance(sheets, list)
            else {}
        )

        return (
            "Google Sheets spreadsheet created:\n"
            f"Title: {(data.get('properties', {}) or {}).get('title', params.title)}\n"
            f"Spreadsheet ID: {spreadsheet_id}\n"
            f"Spreadsheet URL: {spreadsheet_url}\n"
            f"Initial Tab: {first_sheet_props.get('title', params.sheet_title)}\n"
            f"Initial Tab ID: {first_sheet_props.get('sheetId', '-')}"
        )
    except Exception as exc:
        return f"Error creating Google Sheets spreadsheet: {str(exc)}"


@mcp.tool()
async def add_sheet_tab(
    spreadsheet_id: str,
    title: str,
    row_count: int = 1000,
    column_count: int = 26,
) -> str:
    """Adds a new tab (sheet) into an existing spreadsheet."""
    try:
        params = _AddSheetTabInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "title": title,
                "row_count": row_count,
                "column_count": column_count,
            }
        )
        data, err = await _sheets_post(
            f"/{params.spreadsheet_id}:batchUpdate",
            json_body={
                "requests": [
                    {
                        "addSheet": {
                            "properties": {
                                "title": params.title,
                                "gridProperties": {
                                    "rowCount": params.row_count,
                                    "columnCount": params.column_count,
                                },
                            }
                        }
                    }
                ]
            },
        )
        if err:
            return err

        replies = data.get("replies", []) if isinstance(data, dict) else []
        added_props = {}
        if replies and isinstance(replies[0], dict):
            added_props = (replies[0].get("addSheet", {}) or {}).get("properties", {}) or {}

        meta, meta_err = await _sheets_get(
            f"/{params.spreadsheet_id}",
            params={"fields": "spreadsheetUrl,properties.title"},
        )
        if meta_err:
            meta = {}

        return (
            "Google Sheets tab created:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Spreadsheet Title: {(meta.get('properties', {}) or {}).get('title', '-')}\n"
            f"Tab Title: {added_props.get('title', params.title)}\n"
            f"Tab ID: {added_props.get('sheetId', '-')}\n"
            f"Grid: {params.row_count}x{params.column_count}\n"
            f"Spreadsheet URL: {meta.get('spreadsheetUrl', '-')}"
        )
    except Exception as exc:
        return f"Error creating Google Sheets tab: {str(exc)}"


@mcp.tool()
async def get_spreadsheet_metadata(file_id: str) -> str:
    """Gets metadata for a spreadsheet file (Google Sheets or .xlsx/.xls)."""
    try:
        params = _SpreadsheetFileIdInput.model_validate({"file_id": file_id})
        drive_data, drive_err = await _drive_get(
            f"/files/{params.file_id}",
            params={
                "fields": (
                    "id,name,mimeType,modifiedTime,size,webViewLink,"
                    "owners(displayName,emailAddress),parents"
                ),
                "supportsAllDrives": "true",
            },
        )
        if drive_err:
            return drive_err
        if not drive_data:
            return "Spreadsheet file metadata not found."

        mime_type = str(drive_data.get("mimeType", "")).strip()
        owners = drive_data.get("owners", []) if isinstance(drive_data, dict) else []
        owners_text = ", ".join(
            [
                f"{owner.get('displayName', '-') } <{owner.get('emailAddress', '-')}>"
                for owner in owners
                if isinstance(owner, dict)
            ]
        ) or "-"

        tabs_line = "- (unknown)"
        total_tabs = "-"
        kind = _mime_label(mime_type)

        if mime_type == GOOGLE_SHEET_MIME:
            sheet_data, sheet_err = await _sheets_get(
                f"/{params.file_id}",
                params={"fields": "sheets(properties.sheetId,properties.title)"},
            )
            if sheet_err:
                tabs_line = f"- (warning: {sheet_err})"
            else:
                tabs = sheet_data.get("sheets", []) if isinstance(sheet_data, dict) else []
                total_tabs = str(len(tabs))
                lines = []
                for item in tabs:
                    props = item.get("properties", {}) if isinstance(item, dict) else {}
                    lines.append(f"- {props.get('title', '-')} | Sheet ID: {props.get('sheetId', '-')}")
                tabs_line = "\n".join(lines) if lines else "- (none)"
        elif mime_type in EXCEL_MIMES:
            payload, bytes_err = await _drive_get_bytes(
                f"/files/{params.file_id}",
                params={"alt": "media", "supportsAllDrives": "true"},
            )
            if bytes_err or payload is None:
                tabs_line = f"- (warning: {bytes_err or 'unable to read Excel bytes'})"
            elif mime_type == EXCEL_MIME:
                try:
                    from openpyxl import load_workbook
                except Exception:
                    tabs_line = "- (install openpyxl to inspect .xlsx tabs)"
                else:
                    wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
                    total_tabs = str(len(wb.sheetnames))
                    tabs_line = "\n".join([f"- {name}" for name in wb.sheetnames]) or "- (none)"
            else:
                try:
                    import xlrd
                except Exception:
                    tabs_line = "- (install xlrd to inspect .xls tabs)"
                else:
                    wb = xlrd.open_workbook(file_contents=payload)
                    names = wb.sheet_names()
                    total_tabs = str(len(names))
                    tabs_line = "\n".join([f"- {name}" for name in names]) or "- (none)"

        return (
            "Spreadsheet Metadata:\n"
            f"File ID: {drive_data.get('id', params.file_id)}\n"
            f"Name: {drive_data.get('name', '-')}\n"
            f"Type: {kind}\n"
            f"MIME Type: {mime_type or '-'}\n"
            f"Modified: {drive_data.get('modifiedTime', '-')}\n"
            f"Size: {drive_data.get('size', '-')}\n"
            f"Owners: {owners_text}\n"
            f"Link: {drive_data.get('webViewLink', '-')}\n"
            f"Total Tabs: {total_tabs}\n"
            f"Tabs:\n{tabs_line}"
        )
    except Exception as exc:
        return f"Error getting spreadsheet metadata: {str(exc)}"


@mcp.tool()
async def read_excel_values(
    file_id: str,
    sheet_name: str = "",
    range_a1: str = "A1:Z50",
    max_rows: int = 50,
    max_cols: int = 20,
) -> str:
    """Reads values from a .xlsx/.xls file in Drive without converting it."""
    try:
        params = _ReadExcelValuesInput.model_validate(
            {
                "file_id": file_id,
                "sheet_name": sheet_name,
                "range_a1": range_a1,
                "max_rows": max_rows,
                "max_cols": max_cols,
            }
        )
        meta, meta_err = await _drive_get(
            f"/files/{params.file_id}",
            params={"fields": "id,name,mimeType,webViewLink", "supportsAllDrives": "true"},
        )
        if meta_err:
            return meta_err
        if not meta:
            return "Spreadsheet file not found."

        mime_type = str(meta.get("mimeType", "")).strip()
        if mime_type not in EXCEL_MIMES:
            if mime_type == GOOGLE_SHEET_MIME:
                return (
                    "File is a native Google Sheet. "
                    "Use read_sheet_values(spreadsheet_id=...) for this file."
                )
            return (
                f"File MIME type '{mime_type or '-'}' is not supported by read_excel_values. "
                f"Expected one of: {EXCEL_MIME}, {EXCEL_LEGACY_MIME}."
            )

        payload, bytes_err = await _drive_get_bytes(
            f"/files/{params.file_id}",
            params={"alt": "media", "supportsAllDrives": "true"},
        )
        if bytes_err:
            return bytes_err
        if payload is None:
            return "Failed to download Excel file bytes."

        range_sheet, body_range = _split_sheet_range(params.range_a1)
        file_kind = ".xlsx" if mime_type == EXCEL_MIME else ".xls"

        if mime_type == EXCEL_MIME:
            try:
                from openpyxl import load_workbook
            except Exception:
                return (
                    "openpyxl is required to read .xlsx files directly. "
                    "Install dependency and retry."
                )
            wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
            available_sheets = list(wb.sheetnames)
            target_sheet = params.sheet_name.strip() or range_sheet or (available_sheets[0] if available_sheets else "")
            if not target_sheet:
                return "No worksheet found in the Excel file."
            if target_sheet not in available_sheets:
                return (
                    f"Worksheet '{target_sheet}' not found. "
                    f"Available sheets: {', '.join(available_sheets)}"
                )
            ws = wb[target_sheet]
            reader_mode = "openpyxl"
        else:
            try:
                import xlrd
            except Exception:
                return (
                    "xlrd is required to read .xls files directly. "
                    "Install dependency and retry."
                )
            wb = xlrd.open_workbook(file_contents=payload)
            available_sheets = list(wb.sheet_names())
            target_sheet = params.sheet_name.strip() or range_sheet or (available_sheets[0] if available_sheets else "")
            if not target_sheet:
                return "No worksheet found in the Excel file."
            if target_sheet not in available_sheets:
                return (
                    f"Worksheet '{target_sheet}' not found. "
                    f"Available sheets: {', '.join(available_sheets)}"
                )
            ws = wb.sheet_by_name(target_sheet)
            reader_mode = "xlrd"

        min_col, min_row, max_col, max_row = _parse_excel_bounds(body_range)
        shown_max_row = min(max_row, min_row + params.max_rows - 1)
        shown_max_col = min(max_col, min_col + params.max_cols - 1)

        values = []
        if reader_mode == "openpyxl":
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=shown_max_row,
                min_col=min_col,
                max_col=shown_max_col,
                values_only=True,
            ):
                values.append([_excel_cell_to_text(cell) for cell in row])
        else:
            for row_idx in range(min_row - 1, shown_max_row):
                row_values = []
                for col_idx in range(min_col - 1, shown_max_col):
                    if row_idx < ws.nrows and col_idx < ws.ncols:
                        row_values.append(_excel_cell_to_text(ws.cell_value(row_idx, col_idx)))
                    else:
                        row_values.append("")
                values.append(row_values)

        if not values:
            return (
                f"No values found for file '{params.file_id}' sheet '{target_sheet}' "
                f"range '{body_range}'."
            )

        lines = []
        for offset, row in enumerate(values, start=0):
            lines.append(f"- R{min_row + offset}: {' | '.join(row)}")

        notes = []
        if max_row > shown_max_row:
            notes.append(f"Row display limited to first {params.max_rows} rows.")
        if max_col > shown_max_col:
            notes.append(f"Column display limited to first {params.max_cols} columns.")
        notes_text = ("\nNotes:\n- " + "\n- ".join(notes)) if notes else ""

        return (
            f"Excel ({file_kind}) Values:\n"
            f"File ID: {params.file_id}\n"
            f"File Name: {meta.get('name', '-')}\n"
            f"Sheet: {target_sheet}\n"
            f"Range: {body_range}\n"
            f"Rows Shown: {len(values)}\n"
            f"Link: {meta.get('webViewLink', '-')}\n"
            f"Values:\n{chr(10).join(lines)}{notes_text}"
        )
    except Exception as exc:
        return f"Error reading Excel values: {str(exc)}"


@mcp.tool()
async def convert_excel_to_google_sheet(
    file_id: str,
    new_title: str = "",
    move_to_parent: bool = True,
) -> str:
    """Converts an existing .xlsx/.xls file in Drive into a native Google Sheet."""
    try:
        params = _ConvertExcelInput.model_validate(
            {"file_id": file_id, "new_title": new_title, "move_to_parent": move_to_parent}
        )
        src_meta, src_err = await _drive_get(
            f"/files/{params.file_id}",
            params={
                "fields": "id,name,mimeType,parents,webViewLink",
                "supportsAllDrives": "true",
            },
        )
        if src_err:
            return src_err
        if not src_meta:
            return "Source file not found."

        source_mime = str(src_meta.get("mimeType", "")).strip()
        if source_mime == GOOGLE_SHEET_MIME:
            return (
                "File is already a native Google Sheet.\n"
                f"Spreadsheet ID: {src_meta.get('id', params.file_id)}\n"
                f"Link: {src_meta.get('webViewLink', '-')}"
            )
        if source_mime not in EXCEL_MIMES:
            return (
                f"Source file MIME '{source_mime or '-'}' is not supported for conversion. "
                f"Expected one of: {EXCEL_MIME}, {EXCEL_LEGACY_MIME}."
            )

        source_name = str(src_meta.get("name", "Untitled")).strip() or "Untitled"
        title = params.new_title.strip()
        if not title:
            source_name_lower = source_name.lower()
            if source_name_lower.endswith(".xlsx"):
                title = source_name[:-5]
            elif source_name_lower.endswith(".xls"):
                title = source_name[:-4]
            else:
                title = source_name

        parents = src_meta.get("parents", []) if isinstance(src_meta, dict) else []
        copy_body: dict = {"name": title, "mimeType": GOOGLE_SHEET_MIME}
        if params.move_to_parent and isinstance(parents, list) and parents:
            copy_body["parents"] = parents

        copied, copy_err = await _drive_post(
            f"/files/{params.file_id}/copy",
            params={"supportsAllDrives": "true"},
            json_body=copy_body,
        )
        if not copy_err and copied and copied.get("id"):
            new_id = copied.get("id")
            return (
                "Excel to Google Sheets conversion completed:\n"
                f"Source File ID: {params.file_id}\n"
                f"Source Name: {source_name}\n"
                f"Spreadsheet ID: {new_id}\n"
                f"Spreadsheet Name: {copied.get('name', title)}\n"
                f"Link: https://docs.google.com/spreadsheets/d/{new_id}/edit"
            )

        payload, bytes_err = await _drive_get_bytes(
            f"/files/{params.file_id}",
            params={"alt": "media", "supportsAllDrives": "true"},
        )
        if bytes_err:
            return (
                "Failed to convert via Drive copy and failed to download source bytes.\n"
                f"Copy Error: {copy_err or '-'}\n"
                f"Download Error: {bytes_err}"
            )
        if payload is None:
            return f"Failed to convert file: {copy_err or 'unknown error'}"

        upload_metadata: dict = {"name": title, "mimeType": GOOGLE_SHEET_MIME}
        if params.move_to_parent and isinstance(parents, list) and parents:
            upload_metadata["parents"] = parents

        uploaded, upload_err = await _drive_upload_multipart(
            metadata=upload_metadata,
            media_bytes=payload,
            media_mime=source_mime,
            media_filename="source.xlsx" if source_mime == EXCEL_MIME else "source.xls",
        )
        if upload_err:
            return (
                "Failed to convert via Drive copy and multipart upload fallback.\n"
                f"Copy Error: {copy_err or '-'}\n"
                f"Upload Error: {upload_err}"
            )

        new_id = uploaded.get("id", "-") if isinstance(uploaded, dict) else "-"
        return (
            "Excel to Google Sheets conversion completed (upload fallback):\n"
            f"Source File ID: {params.file_id}\n"
            f"Source Name: {source_name}\n"
            f"Spreadsheet ID: {new_id}\n"
            f"Spreadsheet Name: {(uploaded or {}).get('name', title)}\n"
            f"Link: https://docs.google.com/spreadsheets/d/{new_id}/edit"
        )
    except Exception as exc:
        return f"Error converting Excel to Google Sheets: {str(exc)}"


@mcp.tool()
async def export_google_sheet(
    file_id: str,
    export_format: str = "xlsx",
    gid: str | None = None,
    range_a1: str = "",
    max_preview_chars: int = 2000,
) -> str:
    """Exports a native Google Sheet into a supported format."""
    try:
        params = _ExportGoogleSheetInput.model_validate(
            {
                "file_id": file_id,
                "export_format": export_format,
                "gid": gid,
                "range_a1": range_a1,
                "max_preview_chars": max_preview_chars,
            }
        )
        meta, meta_err = await _drive_get(
            f"/files/{params.file_id}",
            params={"fields": "id,name,mimeType,webViewLink", "supportsAllDrives": "true"},
        )
        if meta_err:
            return meta_err
        if not meta:
            return "Spreadsheet file not found."

        mime_type = str(meta.get("mimeType", "")).strip()
        if mime_type != GOOGLE_SHEET_MIME:
            if mime_type in EXCEL_MIMES:
                return (
                    "File is an Excel spreadsheet (.xlsx/.xls), not a native Google Sheet. "
                    "Use convert_excel_to_google_sheet(file_id=...) first, then export."
                )
            return (
                f"File MIME type '{mime_type or '-'}' is not exportable by export_google_sheet. "
                f"Expected '{GOOGLE_SHEET_MIME}'."
            )

        export_meta = SHEET_EXPORT_FORMATS[params.export_format]
        payload, export_err = await _drive_export_google_sheet_bytes(
            file_id=params.file_id,
            export_format=params.export_format,
            mime_type=str(export_meta["mime_type"]),
            gid=params.gid,
            range_a1=params.range_a1,
        )
        if export_err:
            return export_err
        if payload is None:
            return "Failed to export Google Sheet."

        size = len(payload)
        is_binary = bool(export_meta["binary"])
        base = (
            "Google Sheets export completed:\n"
            f"Spreadsheet ID: {params.file_id}\n"
            f"Spreadsheet Name: {meta.get('name', '-')}\n"
            f"Format: {params.export_format}\n"
            f"Export MIME: {export_meta['mime_type']}\n"
            f"Byte Size: {size}\n"
            f"Source Link: {meta.get('webViewLink', '-')}"
        )

        if is_binary:
            return (
                f"{base}\n"
                "Preview: (binary export, preview omitted).\n"
                "Tip: use format=csv or format=tsv for text preview."
            )

        decoded = _decode_text_with_fallback(payload)
        preview = decoded[: params.max_preview_chars]
        truncated = "yes" if len(decoded) > len(preview) else "no"
        return (
            f"{base}\n"
            f"Preview Truncated: {truncated}\n"
            f"Preview:\n{preview}"
        )
    except Exception as exc:
        return f"Error exporting Google Sheet: {str(exc)}"


@mcp.tool()
async def batch_get_sheet_values(
    spreadsheet_id: str,
    ranges: list[str],
    value_render_option: str = "FORMATTED_VALUE",
    date_time_render_option: str = "SERIAL_NUMBER",
    max_rows_per_range: int = 20,
    max_cols_per_row: int = 20,
) -> str:
    """Reads multiple ranges in one Sheets API batchGet request."""
    try:
        params = _BatchGetSheetValuesInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "ranges": ranges,
                "value_render_option": value_render_option,
                "date_time_render_option": date_time_render_option,
                "max_rows_per_range": max_rows_per_range,
                "max_cols_per_row": max_cols_per_row,
            }
        )
        data, err = await _sheets_get(
            f"/{params.spreadsheet_id}/values:batchGet",
            params={
                "ranges": params.ranges,
                "valueRenderOption": params.value_render_option,
                "dateTimeRenderOption": params.date_time_render_option,
            },
        )
        if err:
            return err
        if not data:
            return "No value response from Google Sheets batchGet."

        value_ranges = data.get("valueRanges", []) if isinstance(data, dict) else []
        if not value_ranges:
            return "No values returned for the requested ranges."

        sections: list[str] = []
        for item in value_ranges:
            if not isinstance(item, dict):
                continue
            returned_range = str(item.get("range", "-"))
            values = item.get("values", [])
            if not isinstance(values, list):
                values = []
            row_count = len(values)
            shown_rows = values[: params.max_rows_per_range]
            row_lines: list[str] = []
            cols_truncated = False
            for idx, row in enumerate(shown_rows, start=1):
                if not isinstance(row, list):
                    row = [row]
                normalized_row = [str(cell) for cell in row[: params.max_cols_per_row]]
                if len(row) > params.max_cols_per_row:
                    cols_truncated = True
                row_lines.append(f"  - R{idx}: {' | '.join(normalized_row)}")
            notes: list[str] = []
            if row_count > len(shown_rows):
                notes.append(f"rows limited to first {params.max_rows_per_range}")
            if cols_truncated:
                notes.append(f"columns limited to first {params.max_cols_per_row}")
            note_suffix = f" ({'; '.join(notes)})" if notes else ""
            section = (
                f"- Range: {returned_range}\n"
                f"  Rows Returned: {row_count}\n"
                f"  Values:\n"
                + ("\n".join(row_lines) if row_lines else "  - (empty)")
                + note_suffix
            )
            sections.append(section)

        return (
            "Google Sheets batch get values completed:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Ranges Requested: {len(params.ranges)}\n"
            f"Ranges Returned: {len(sections)}\n"
            f"Value Render Option: {params.value_render_option}\n"
            f"Date Time Render Option: {params.date_time_render_option}\n"
            f"Results:\n{chr(10).join(sections)}"
        )
    except Exception as exc:
        return f"Error reading Google Sheets batch values: {str(exc)}"


@mcp.tool()
async def batch_update_sheet_values(
    spreadsheet_id: str,
    updates: list[dict],
    value_input_option: str = "USER_ENTERED",
) -> str:
    """Updates multiple ranges in one Sheets API batchUpdate request."""
    try:
        params = _BatchUpdateSheetValuesInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "updates": updates,
                "value_input_option": value_input_option,
            }
        )
        data_payload = []
        for update in params.updates:
            normalized_values = [
                [str(cell) if cell is not None else "" for cell in row]
                for row in update.values
            ]
            data_payload.append(
                {
                    "range": update.range_a1,
                    "majorDimension": "ROWS",
                    "values": normalized_values,
                }
            )

        data, err = await _sheets_post(
            f"/{params.spreadsheet_id}/values:batchUpdate",
            json_body={
                "valueInputOption": params.value_input_option,
                "data": data_payload,
                "includeValuesInResponse": False,
            },
        )
        if err:
            return err
        if not data:
            return "No update response from Google Sheets batchUpdate."

        responses = data.get("responses", []) if isinstance(data, dict) else []
        response_lines: list[str] = []
        for idx, resp in enumerate(responses, start=1):
            if not isinstance(resp, dict):
                continue
            response_lines.append(
                f"- Update {idx}: {resp.get('updatedRange', '-')}"
                f" | Rows: {resp.get('updatedRows', '-')}"
                f" | Cols: {resp.get('updatedColumns', '-')}"
                f" | Cells: {resp.get('updatedCells', '-')}"
            )

        return (
            "Google Sheets batch update completed:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Updates Requested: {len(params.updates)}\n"
            f"Updated Ranges: {len(responses)}\n"
            f"Total Updated Rows: {data.get('totalUpdatedRows', '-')}\n"
            f"Total Updated Columns: {data.get('totalUpdatedColumns', '-')}\n"
            f"Total Updated Cells: {data.get('totalUpdatedCells', '-')}\n"
            f"Value Input Option: {params.value_input_option}\n"
            f"Response Details:\n{chr(10).join(response_lines) if response_lines else '- (none)'}"
        )
    except Exception as exc:
        return f"Error updating Google Sheets batch values: {str(exc)}"


@mcp.tool()
async def share_spreadsheet(
    file_id: str,
    user_email: str,
    role: str = "writer",
    send_notification: bool = True,
    message: str = "",
) -> str:
    """Shares a spreadsheet file (native Sheets or .xlsx/.xls) with one user."""
    try:
        params = _ShareSpreadsheetInput.model_validate(
            {
                "file_id": file_id,
                "user_email": user_email,
                "role": role,
                "send_notification": send_notification,
                "message": message,
            }
        )
        meta, meta_err = await _drive_get(
            f"/files/{params.file_id}",
            params={"fields": "id,name,mimeType,webViewLink", "supportsAllDrives": "true"},
        )
        if meta_err:
            return meta_err
        if not meta:
            return "Spreadsheet file not found."

        mime_type = str(meta.get("mimeType", "")).strip()
        if mime_type not in ({GOOGLE_SHEET_MIME} | EXCEL_MIMES):
            return (
                f"File MIME type '{mime_type or '-'}' is not a spreadsheet type supported by "
                "share_spreadsheet."
            )

        request_params = {
            "supportsAllDrives": "true",
            "sendNotificationEmail": "true" if params.send_notification else "false",
        }
        if params.send_notification and params.message:
            request_params["emailMessage"] = params.message

        perm, perm_err = await _drive_post(
            f"/files/{params.file_id}/permissions",
            params=request_params,
            json_body={
                "type": "user",
                "role": params.role,
                "emailAddress": params.user_email,
            },
        )
        if perm_err:
            return perm_err

        return (
            "Spreadsheet share completed:\n"
            f"File ID: {params.file_id}\n"
            f"File Name: {meta.get('name', '-')}\n"
            f"Type: {_mime_label(mime_type)}\n"
            f"Shared With: {params.user_email}\n"
            f"Role: {params.role}\n"
            f"Notification Sent: {params.send_notification}\n"
            f"Permission ID: {(perm or {}).get('id', '-')}\n"
            f"Link: {meta.get('webViewLink', '-')}"
        )
    except Exception as exc:
        return f"Error sharing spreadsheet: {str(exc)}"


@mcp.tool()
async def create_spreadsheet_from_template(
    template_file_id: str,
    new_title: str,
    destination_folder_id: str = "",
) -> str:
    """Creates a new spreadsheet by copying an existing spreadsheet template file."""
    try:
        params = _CreateSpreadsheetFromTemplateInput.model_validate(
            {
                "template_file_id": template_file_id,
                "new_title": new_title,
                "destination_folder_id": destination_folder_id,
            }
        )
        source_meta, source_err = await _drive_get(
            f"/files/{params.template_file_id}",
            params={
                "fields": "id,name,mimeType,parents,webViewLink",
                "supportsAllDrives": "true",
            },
        )
        if source_err:
            return source_err
        if not source_meta:
            return "Template spreadsheet file not found."

        source_mime = str(source_meta.get("mimeType", "")).strip()
        if source_mime not in ({GOOGLE_SHEET_MIME} | EXCEL_MIMES):
            return (
                f"Template MIME type '{source_mime or '-'}' is not supported. "
                "Expected native Google Sheet or .xlsx/.xls file."
            )

        copy_body: dict[str, Any] = {"name": params.new_title}
        destination_folder = params.destination_folder_id.strip()
        if destination_folder:
            copy_body["parents"] = [destination_folder]
        else:
            parents = source_meta.get("parents", []) if isinstance(source_meta, dict) else []
            if isinstance(parents, list) and parents:
                copy_body["parents"] = parents

        copied, copy_err = await _drive_post(
            f"/files/{params.template_file_id}/copy",
            params={"supportsAllDrives": "true"},
            json_body=copy_body,
        )
        if copy_err:
            return copy_err
        if not copied:
            return "Failed to copy spreadsheet template."

        new_id = str(copied.get("id", "")).strip()
        if not new_id:
            return "Template copy completed but returned no file id."

        copied_meta, copied_meta_err = await _drive_get(
            f"/files/{new_id}",
            params={
                "fields": "id,name,mimeType,webViewLink,parents",
                "supportsAllDrives": "true",
            },
        )
        if copied_meta_err:
            copied_meta = copied or {}

        return (
            "Spreadsheet created from template:\n"
            f"Template File ID: {params.template_file_id}\n"
            f"Template Name: {source_meta.get('name', '-')}\n"
            f"Source Type: {_mime_label(source_mime)}\n"
            f"New File ID: {(copied_meta or {}).get('id', new_id)}\n"
            f"New File Name: {(copied_meta or {}).get('name', params.new_title)}\n"
            f"New File Type: {_mime_label(str((copied_meta or {}).get('mimeType', source_mime)).strip())}\n"
            f"Destination Folder ID: {destination_folder or '-'}\n"
            f"Link: {(copied_meta or {}).get('webViewLink', f'https://docs.google.com/spreadsheets/d/{new_id}/edit')}"
        )
    except Exception as exc:
        return f"Error creating spreadsheet from template: {str(exc)}"


@mcp.tool()
async def import_csv_to_sheet(
    spreadsheet_id: str,
    sheet_name: str,
    csv_text: str,
    overwrite: bool = False,
) -> str:
    """Imports CSV text into a target sheet tab."""
    try:
        params = _ImportCsvToSheetInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "sheet_name": sheet_name,
                "csv_text": csv_text,
                "overwrite": overwrite,
            }
        )
        parsed_rows = [row for row in csv.reader(StringIO(params.csv_text))]
        if not parsed_rows:
            return "CSV input is empty."

        normalized_rows = [[str(cell) for cell in row] for row in parsed_rows]
        quoted_sheet_name = _quote_sheet_name(params.sheet_name)
        target_start = f"{quoted_sheet_name}!A1"

        if params.overwrite:
            clear_range = f"{quoted_sheet_name}!A:ZZZ"
            _, clear_err = await _sheets_post(
                f"/{params.spreadsheet_id}/values/{_encode_a1_range(clear_range)}:clear",
                json_body={},
            )
            if clear_err:
                return clear_err

        data, err = await _sheets_put(
            f"/{params.spreadsheet_id}/values/{_encode_a1_range(target_start)}",
            params={"valueInputOption": "USER_ENTERED"},
            json_body={"majorDimension": "ROWS", "values": normalized_rows},
        )
        if err:
            return err

        return (
            "CSV import to sheet completed:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Sheet: {params.sheet_name}\n"
            f"Overwrite: {params.overwrite}\n"
            f"Rows Imported: {len(normalized_rows)}\n"
            f"Updated Range: {(data or {}).get('updatedRange', '-')}\n"
            f"Updated Rows: {(data or {}).get('updatedRows', '-')}\n"
            f"Updated Columns: {(data or {}).get('updatedColumns', '-')}\n"
            f"Updated Cells: {(data or {}).get('updatedCells', '-')}"
        )
    except Exception as exc:
        return f"Error importing CSV to sheet: {str(exc)}"


@mcp.tool()
async def insert_sheet_chart(
    spreadsheet_id: str,
    sheet_id: int,
    chart_spec: dict[str, Any],
) -> str:
    """Inserts a chart using batchUpdate addChart request."""
    try:
        params = _InsertSheetChartInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "sheet_id": sheet_id,
                "chart_spec": chart_spec,
            }
        )
        chart_payload = dict(params.chart_spec)
        if "spec" not in chart_payload:
            chart_payload = {"spec": dict(params.chart_spec)}
        if "position" not in chart_payload:
            chart_payload["position"] = {
                "overlayPosition": {
                    "anchorCell": {
                        "sheetId": params.sheet_id,
                        "rowIndex": 0,
                        "columnIndex": 0,
                    }
                }
            }

        data, err = await _sheets_post(
            f"/{params.spreadsheet_id}:batchUpdate",
            json_body={"requests": [{"addChart": {"chart": chart_payload}}]},
        )
        if err:
            return err

        replies = data.get("replies", []) if isinstance(data, dict) else []
        chart_id = "-"
        if replies and isinstance(replies[0], dict):
            chart_id = (
                ((replies[0].get("addChart", {}) or {}).get("chart", {}) or {}).get("chartId", "-")
            )

        return (
            "Sheet chart inserted:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Sheet ID: {params.sheet_id}\n"
            f"Chart ID: {chart_id}"
        )
    except Exception as exc:
        return f"Error inserting sheet chart: {str(exc)}"


@mcp.tool()
async def protect_sheet_or_range(
    spreadsheet_id: str,
    sheet_id: int | None = None,
    range_a1: str = "",
    editors: list[str] | None = None,
    warning_only: bool = False,
) -> str:
    """Protects a full sheet or a specific A1 range in a spreadsheet."""
    try:
        params = _ProtectSheetOrRangeInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "sheet_id": sheet_id,
                "range_a1": range_a1,
                "editors": editors or [],
                "warning_only": warning_only,
            }
        )

        target_range: dict[str, Any] | None = None
        target_sheet_name = "-"
        if params.range_a1:
            resolved_grid, resolved_name, resolve_err = await _resolve_grid_range_from_a1(
                params.spreadsheet_id,
                params.range_a1,
                fallback_sheet_id=params.sheet_id,
            )
            if resolve_err:
                return resolve_err
            target_range = resolved_grid
            target_sheet_name = resolved_name or "-"
        else:
            _, id_to_title, _, map_err = await _get_sheet_mappings(params.spreadsheet_id)
            if map_err:
                return map_err
            if params.sheet_id is None:
                return "sheet_id is required when range_a1 is empty."
            target_sheet_name = id_to_title.get(params.sheet_id, "-")
            target_range = {"sheetId": params.sheet_id}

        protected_range: dict[str, Any] = {
            "range": target_range,
            "warningOnly": params.warning_only,
        }
        if params.editors:
            protected_range["editors"] = {"users": params.editors}

        data, err = await _sheets_post(
            f"/{params.spreadsheet_id}:batchUpdate",
            json_body={
                "requests": [
                    {
                        "addProtectedRange": {
                            "protectedRange": protected_range,
                        }
                    }
                ]
            },
        )
        if err:
            return err

        replies = data.get("replies", []) if isinstance(data, dict) else []
        protected_range_id = "-"
        if replies and isinstance(replies[0], dict):
            protected_range_id = (
                ((replies[0].get("addProtectedRange", {}) or {}).get("protectedRange", {}) or {}).get(
                    "protectedRangeId",
                    "-",
                )
            )

        scope = params.range_a1 if params.range_a1 else f"Sheet ID {params.sheet_id}"
        return (
            "Sheet protection created:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Scope: {scope}\n"
            f"Sheet Name: {target_sheet_name}\n"
            f"Warning Only: {params.warning_only}\n"
            f"Editors: {', '.join(params.editors) if params.editors else '-'}\n"
            f"Protected Range ID: {protected_range_id}"
        )
    except Exception as exc:
        return f"Error creating sheet protection: {str(exc)}"


@mcp.tool()
async def create_pivot_table(
    spreadsheet_id: str,
    source_range: str,
    target_sheet: str,
    target_cell: str = "A1",
    summarize_function: str = "COUNTA",
) -> str:
    """Creates a basic pivot table at target cell using source range."""
    try:
        params = _CreatePivotTableInput.model_validate(
            {
                "spreadsheet_id": spreadsheet_id,
                "source_range": source_range,
                "target_sheet": target_sheet,
                "target_cell": target_cell,
                "summarize_function": summarize_function,
            }
        )
        title_to_id, _, spreadsheet_url, map_err = await _get_sheet_mappings(params.spreadsheet_id)
        if map_err:
            return map_err
        target_sheet_id = title_to_id.get(params.target_sheet.lower())
        if target_sheet_id is None:
            return f"Target sheet '{params.target_sheet}' not found in spreadsheet."

        source_grid, _, source_err = await _resolve_grid_range_from_a1(
            params.spreadsheet_id,
            params.source_range,
            fallback_sheet_id=target_sheet_id,
        )
        if source_err:
            return source_err
        if not source_grid:
            return "Failed to resolve source range."

        start_row, start_col, cell_err = _a1_cell_to_index(params.target_cell)
        if cell_err:
            return cell_err
        if start_row is None or start_col is None:
            return "Failed to resolve target cell."

        source_width = int(source_grid.get("endColumnIndex", 0)) - int(
            source_grid.get("startColumnIndex", 0)
        )
        row_offset = 0
        value_offset = 1 if source_width > 1 else 0

        pivot_table = {
            "source": source_grid,
            "rows": [
                {
                    "sourceColumnOffset": row_offset,
                    "showTotals": True,
                    "sortOrder": "ASCENDING",
                }
            ],
            "values": [
                {
                    "summarizeFunction": params.summarize_function,
                    "sourceColumnOffset": value_offset,
                    "name": f"{params.summarize_function} (column {value_offset + 1})",
                }
            ],
        }

        _, err = await _sheets_post(
            f"/{params.spreadsheet_id}:batchUpdate",
            json_body={
                "requests": [
                    {
                        "updateCells": {
                            "start": {
                                "sheetId": target_sheet_id,
                                "rowIndex": start_row,
                                "columnIndex": start_col,
                            },
                            "rows": [{"values": [{"pivotTable": pivot_table}]}],
                            "fields": "pivotTable",
                        }
                    }
                ]
            },
        )
        if err:
            return err

        return (
            "Pivot table created:\n"
            f"Spreadsheet ID: {params.spreadsheet_id}\n"
            f"Source Range: {params.source_range}\n"
            f"Target Sheet: {params.target_sheet}\n"
            f"Target Cell: {params.target_cell}\n"
            f"Summarize Function: {params.summarize_function}\n"
            f"Spreadsheet URL: {spreadsheet_url}"
        )
    except Exception as exc:
        return f"Error creating pivot table: {str(exc)}"


@mcp.tool()
async def get_spreadsheet_permissions(file_id: str) -> str:
    """Lists Drive permissions for a spreadsheet file."""
    try:
        params = _SpreadsheetFileIdInput.model_validate({"file_id": file_id})
        meta, meta_err = await _drive_get(
            f"/files/{params.file_id}",
            params={"fields": "id,name,mimeType,webViewLink", "supportsAllDrives": "true"},
        )
        if meta_err:
            return meta_err
        if not meta:
            return "Spreadsheet file not found."

        mime_type = str(meta.get("mimeType", "")).strip()
        if mime_type not in ({GOOGLE_SHEET_MIME} | EXCEL_MIMES):
            return (
                f"File MIME type '{mime_type or '-'}' is not a spreadsheet type supported by "
                "get_spreadsheet_permissions."
            )

        permissions_data, permissions_err = await _drive_get(
            f"/files/{params.file_id}/permissions",
            params={
                "fields": (
                    "permissions(id,type,role,emailAddress,domain,allowFileDiscovery,displayName),"
                    "nextPageToken"
                ),
                "supportsAllDrives": "true",
            },
        )
        if permissions_err:
            return permissions_err

        permissions = (
            permissions_data.get("permissions", [])
            if isinstance(permissions_data, dict)
            else []
        )
        if not permissions:
            return (
                "Spreadsheet permissions:\n"
                f"File ID: {params.file_id}\n"
                f"File Name: {meta.get('name', '-')}\n"
                "Permissions: (none)"
            )

        lines: list[str] = []
        for item in permissions:
            if not isinstance(item, dict):
                continue
            principal = (
                item.get("emailAddress")
                or item.get("domain")
                or item.get("displayName")
                or "-"
            )
            lines.append(
                f"- ID: {item.get('id', '-')}"
                f" | Type: {item.get('type', '-')}"
                f" | Role: {item.get('role', '-')}"
                f" | Principal: {principal}"
            )

        return (
            "Spreadsheet permissions:\n"
            f"File ID: {params.file_id}\n"
            f"File Name: {meta.get('name', '-')}\n"
            f"Type: {_mime_label(mime_type)}\n"
            f"Link: {meta.get('webViewLink', '-')}\n"
            f"Permission Count: {len(lines)}\n"
            f"Permissions:\n{chr(10).join(lines)}"
        )
    except Exception as exc:
        return f"Error getting spreadsheet permissions: {str(exc)}"


def run() -> None:
    mcp.run()


if __name__ == "__main__":
    run()
